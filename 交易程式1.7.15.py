# 這個版本要檢查三個問題：
# 1. 觸發漲停進場後，沒有在接下來的等待期間加入「等待第 X 分鐘」的訊息。
# 2. 拉高進場和漲停進場同一時間觸發，沒有以漲停進場為優先。
# 3. 領漲替換之後沒有繼續觀察，好像沒有其他動作。
# 這個版本要新增的功能：
# 1. 回測函數的相似度判斷（已完成）
# 2. 交易函數的相似度判斷（已完成）

# -------------------- 事前準備 --------------------
import json
import os
import math
import subprocess
import sys
import time as time_module
import warnings
import msvcrt
import traceback
import shioaji_logic
import importlib
import csv
import threading

from datetime import datetime, time, timedelta, date
from concurrent.futures import ThreadPoolExecutor, as_completed

REQUIRED = [
    ("fugle_marketdata", "fugle-marketdata"),
    ("pandas",           "pandas"),
    ("yaml",             "pyyaml"),
    ("numpy",            "numpy"),
    ("colorama",         "colorama"),
    ("tabulate",         "tabulate"),
    ("openpyxl",         "openpyxl"),
    ("dateutil",         "python-dateutil"),
    ("matplotlib",       "matplotlib")
]

def ensure_packages(pkgs):
    """檢查→缺少就 pip install→最後再動態 import 回來"""
    missing = []
    for mod, pkg in pkgs:
        try:
            importlib.import_module(mod)
        except ImportError:
            missing.append(pkg)

    if missing:
        print("首次執行偵測到以下套件尚未安裝：", ", ".join(missing))
        for pkg in missing:
            subprocess.check_call(
                [sys.executable, "-m", "pip", "install", pkg]
            )
        # 安裝完再把它們 import 進來，程式不用重開
        for mod, pkg in pkgs:
            globals()[mod] = importlib.import_module(mod)
    else:
        print("👍  所有必要套件都已安裝")

ensure_packages(REQUIRED)

import fugle_marketdata as fg
import pandas as pd
import yaml
import numpy as np
import colorama
import shioaji as sj
import touchprice as tp
import requests, bs4
import orjson
import matplotlib.pyplot as plt
from tabulate import tabulate
from openpyxl.styles import PatternFill
from colorama import init, Fore, Style
from fugle_marketdata import RestClient
from bs4 import BeautifulSoup

plt.rcParams['axes.unicode_minus'] = False
colorama.init(autoreset=True)
warnings.filterwarnings("ignore", category=FutureWarning)

# 全域旗標：按下 Q 鍵觸發平倉選單
quit_flag = {"quit": False}

RED = Fore.RED
GREEN = Fore.GREEN
YELLOW = Fore.YELLOW
BLUE = Fore.BLUE
RESET = Style.RESET_ALL

pd.set_option('future.no_silent_downcasting', True)

def _crawl_tw_isin_table(mode: str):
    """
    mode = '2' → 上市股票
    mode = '4' → 上櫃股票
    回傳 [(代號, 中文名), ...]
    """

    url = f"https://isin.twse.com.tw/isin/C_public.jsp?strMode={mode}"
    r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
    r.encoding = "big5"                    # 官方網頁以 Big5 編碼
    soup = bs4.BeautifulSoup(r.text, "lxml")
    rows = soup.select("table tr")[1:]     # 第 0 列是表頭

    pairs = []
    for tr in rows:
        tds = tr.find_all("td")
        if not tds:
            continue
        raw = tds[0].text.strip()
        if raw[:4].isdigit():              # 只要前 4 碼是純數字的股票
            code = raw[:4]
            name = raw.split("\u3000", 1)[1] if "\u3000" in raw else raw[4:]
            pairs.append((code, name))
    return pairs

def fetch_twse_stock_codes(save_json=None, save_csv=None):
    """
    取得台灣上市股票代號與中文名稱清單
    --------------------------------------------------
    Parameters
    ----------
    save_json : str | None
        若給檔名，將結果存成 JSON，例如 "twse_stocks.json"
    save_csv  : str | None
        若給檔名，將結果存成 CSV，例如 "twse_stocks.csv"

    Returns
    -------
    List[Tuple[str,str]]
        [('1101', '台泥'), ('1102', '亞泥'), ...]
    """
    url     = "https://isin.twse.com.tw/isin/C_public.jsp?strMode=2"
    headers = {"User-Agent": "Mozilla/5.0"}

    res = requests.get(url, headers=headers, timeout=10)
    # 網頁採 Big‑5，手動指定編碼避免亂碼
    res.encoding = "big5"

    soup = BeautifulSoup(res.text, "lxml")
    rows = soup.select("table tr")[1:]          # 跳過表頭

    stocks = []
    for r in rows:
        cols = [c.text.strip() for c in r.find_all("td")]
        if not cols:
            continue
        code_name = cols[0]                     # 例：「1101　台泥」
        if len(code_name) >= 4 and code_name[:4].isdigit():
            code = code_name[:4]
            # 以「全形空格」劃分取中文名稱；若切不到就直接取剩餘字串
            name = code_name.split("\u3000", 1)[1] if "\u3000" in code_name else code_name[4:]
            stocks.append((code, name))

    # ----------- (選用) 存檔 -----------
    if save_json:
        with open(save_json, "w", encoding="utf-8") as f:
            json.dump(stocks, f, ensure_ascii=False, indent=2)
    if save_csv:
        with open(save_csv, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Code", "Name"])
            w.writerows(stocks)

    return stocks

STOCK_NAME_MAP = {}      # 全域字典 { "1101": "台泥", ... }

def load_twse_name_map(json_path="twse_stocks_by_market.json"):
    global STOCK_NAME_MAP
    if STOCK_NAME_MAP:
        return

    try:
        if os.path.exists(json_path):
            with open(json_path, "r", encoding="utf-8") as f:
                STOCK_NAME_MAP = json.load(f)
            return

        # 分別抓上市與上櫃
        listed_pairs = _crawl_tw_isin_table("2")  # 上市
        otc_pairs = _crawl_tw_isin_table("4")     # 上櫃

        tse_map = {c: n for c, n in listed_pairs}
        otc_map = {c: n for c, n in otc_pairs}

        STOCK_NAME_MAP = {
            "TSE": tse_map,
            "OTC": otc_map
        }

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(STOCK_NAME_MAP, f, ensure_ascii=False, indent=2)

    except Exception as e:
        print(f"載入股票中文名稱失敗：{e}")
        STOCK_NAME_MAP = {}

def get_stock_name(code):
    """給定 4 碼股票代號，回傳中文名稱；找不到就回空字串"""
    for market in ["TSE", "OTC"]:
        if code in STOCK_NAME_MAP.get(market, {}):
            return STOCK_NAME_MAP[market][code]
    return ""

load_twse_name_map()
'''
# 測試股票代號是否能連結到中文名稱
print(get_stock_name("2330"))   # 台積電  (上市)
print(get_stock_name("5483"))   # 中美晶  (上櫃)
'''

def init_fugle_client():
    try:
        config = load_config("config.yaml")
        client = RestClient(api_key=config['api_key'])
        print("=" * 50)
        print("從 config.yaml 載入 API 金鑰")
        print("=" * 50)
        return client, config['api_key']
    except FileNotFoundError:
        print("錯誤：config.yaml 文件不存在。")
        sys.exit(1)
    except KeyError:
        print("錯誤：config.yaml 中缺少 'api_key'。")
        sys.exit(1)
    except Exception as e:
        print(f"初始化富果API客戶端時發生錯誤：{e}")
        sys.exit(1)

def load_config(config_file):
    try:
        with open(config_file, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    except FileNotFoundError:
        print(f"錯誤：無法找到 {config_file} 文件。")
        sys.exit(1)
    except yaml.YAMLError as e:
        print(f"錯誤：讀取 {config_file} 文件時發生 YAML 錯誤：{e}")
        sys.exit(1)

capital_per_stock = 0
transaction_fee = 0
transaction_discount = 0
trading_tax = 0
below_50 = 0
price_gap_50_to_100 = 0
price_gap_100_to_500 = 0
price_gap_500_to_1000 = 0
price_gap_above_1000 = 0
allow_reentry_after_stop_loss = False

# -------------------- 進入選單 --------------------
def main_menu():
    global capital_per_stock
    load_settings()
    print('\n' + '=' * 50)
    print(f"\n目前股票的單筆投入資本額為{capital_per_stock}萬元")
    while True:
        print("請選擇功能：")
        print("1. 回測程式")
        print("2. 下單程式")
        print("3. 管理族群")
        print("4. 設定選單")
        print("5. 更新K線數據")
        print("6. 查詢處置股")
        print("0. 退出程式")
        print('\n' + '=' * 50)
        choice = input("請輸入選項：")
        if choice == '1':
            backtesting_menu_list()
        elif choice == '2':
            trading_menu_list()
        elif choice == '3':
            manage_groups()
        elif choice == '4':
            settings_menu()
        elif choice == '5':
            update_kline_data_menu()
        elif choice == '6':
            display_disposition_stocks()
        elif choice == '0':
            print("退出程式...再見")
            break
        else:
            print("無效的選項，請重新輸入")

# -------------------- 回測程式 --------------------
def backtesting_menu_list():
    print('\n' + '=' * 50)
    print("\n請選擇功能：")
    print("1. 計算平均過高、2. 自選進場模式、3. 極大化利潤模式、0. 返回主選單")
    print('\n' + '=' * 50)
    back_choice = input("請選擇功能：")
    if back_choice == '1':
        calculate_average_over_high_list()
    elif back_choice == '2':
        simulate_trading_menu()
    elif back_choice == '3':
        maximize_profit_analysis()
    elif back_choice == '0':
        main_menu()
    else:
        print("無效的選項，請重新輸入")

# ------------------ 回測程式：計算平均過高 ------------------
def calculate_average_over_high_list():
    while True:
        print('\n' + '=' * 50)
        print("選擇計算平均過高的模式：")
        print("1. 單一族群分析")
        print("2. 全部族群分析")
        print("0. 返回主選單")
        
        sub_choice = input("請輸入選項：")
        if sub_choice == '1':
            calculate_average_over_high()
        elif sub_choice == '2':
            matrix_dict_analysis = load_matrix_dict_analysis()
            all_group_names = list(matrix_dict_analysis.keys())
            if not all_group_names:
                print("沒有任何族群資料可供分析。")
                continue
            print("開始分析所有族群中的股票...")
            all_group_over_high_averages = []

            for i, group in enumerate(all_group_names):
                print(f"\n=== 分析族群：{group} ===")
                group_average = calculate_average_over_high(group_name=group)
                if group_average is not None:
                    all_group_over_high_averages.append(group_average)
                    
            if all_group_over_high_averages:
                overall_group_average = sum(all_group_over_high_averages) / len(all_group_over_high_averages)
                print(f"\n全部族群的平均過高間隔：{overall_group_average:.2f} 分鐘")
            else:
                print("\n沒有任何族群發生過高間隔的情形。")
        elif sub_choice == '0':
            main_menu()
        else:
            print("無效的選項，請重新輸入")

# ---------------- 回測程式：單一族群分析 ----------------
def calculate_average_over_high(group_name=None):
    daily_kline_data, intraday_kline_data = load_kline_data()

    matrix_dict_analysis = load_matrix_dict_analysis()
    
    if group_name is None:
        group_name = input("請輸入要分析的族群名稱：")
    
    if group_name not in matrix_dict_analysis:
        print("沒有此族群資料")
        return None

    symbols_to_analyze = matrix_dict_analysis[group_name]
    disposition_stocks = load_disposition_stocks()
    symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]

    if not symbols_to_analyze:
        print(f"{group_name} 中沒有可供分析的股票。")
        return None

    print(f"開始分析族群 {group_name} 中的股票...")
    any_condition_one_triggered = False 
    group_over_high_averages = []

    for symbol in symbols_to_analyze:
        print(f"\n正在分析股票：{symbol}")
        
        if symbol not in daily_kline_data or symbol not in intraday_kline_data:
            print(f"無法取得 {symbol} 的日 K 線或一分 K 線數據，跳過。")
            continue
        
        daily_kline_df = pd.DataFrame(daily_kline_data[symbol])
        intraday_data = pd.DataFrame(intraday_kline_data[symbol])

        condition_one_triggered = False
        condition_two_triggered = False
        previous_high = None
        condition_two_time = None
        over_high_intervals = []

        for idx, row in intraday_data.iterrows():
            current_time = pd.to_datetime(row['time']).time()
            if previous_high is None:
                previous_high = row['high']
                continue

            if not condition_one_triggered:
                if row['2min_pct_increase'] >= 2:
                    condition_one_triggered = True
                    condition_two_triggered = False
                    any_condition_one_triggered = True

                    print(f"{symbol} 觸發條件一，開始監測兩分鐘漲幅，兩分鐘漲幅: {row['2min_pct_increase']:.2f}%")

            if condition_one_triggered and not condition_two_triggered:
                if row['high'] <= previous_high:
                    current_time_str = current_time.strftime('%H:%M:%S')
                    print(f"{symbol} 觸發條件二！時間：{current_time_str}")

                    condition_two_time = current_time
                    condition_two_triggered = True

            elif condition_two_triggered:
                if row['highest'] > previous_high:
                    condition_three_time_str = current_time.strftime('%H:%M:%S')
                    print(f"{symbol} 觸發條件三！時間：{condition_three_time_str}")
                    if condition_two_time:
                        today = datetime.today().date()
                        condition_two_datetime = datetime.combine(today, condition_two_time)
                        condition_three_datetime = datetime.combine(today, current_time)
                        interval = (condition_three_datetime - condition_two_datetime).total_seconds() / 60
                        print(f"{symbol} 過高間隔：{interval:.2f} 分鐘")
                        over_high_intervals.append(interval)

                    condition_one_triggered = False
                    condition_two_triggered = False
                    condition_two_time = None

            previous_high = row['high']

        if over_high_intervals:
            q1 = np.percentile(over_high_intervals, 25)
            q3 = np.percentile(over_high_intervals, 75)
            iqr = q3 - q1
            lower_bound = q1 - 1.5 * iqr
            upper_bound = q3 + 1.5 * iqr
            filtered_intervals = [interval for interval in over_high_intervals if lower_bound <= interval <= upper_bound]
            if filtered_intervals:
                average_interval = sum(filtered_intervals) / len(filtered_intervals)
                print(f"{symbol} 平均過高間隔：{average_interval:.2f} 分鐘")
                group_over_high_averages.append(average_interval)
            else:
                print(f"{symbol} 沒有有效的過高間隔數據")
        else:
            print(f"{symbol} 沒有觸發過高間隔的情形")

    if group_over_high_averages:
        group_average_over_high = sum(group_over_high_averages) / len(group_over_high_averages)
        print(f"{group_name} 平均過高間隔：{group_average_over_high:.2f} 分鐘")
        return group_average_over_high
    else:
        print(f"{group_name} 沒有有效的過高間隔數據")
        return None

# ------------------ 回測程式：自選進場模式 ------------------
def simulate_trading_menu():
    matrix_dict_analysis = load_matrix_dict_analysis()
    if not matrix_dict_analysis:
        print("沒有族群資料，請先管理族群。")
        return

    while True:
        print("請選擇操作：")
        print("1. 分析單一族群")
        print("2. 分析全部族群")
        print("0. 返回主選單")
        choice = input("請輸入選項編號：")

        if choice == '1':
            group_name = input("請輸入要分析的族群名稱：")
            if group_name not in matrix_dict_analysis:
                print("沒有此族群資料")
                continue

            try:
                wait_minutes = int(input("請輸入等待時間（分鐘）："))
            except ValueError:
                print("等待時間必須是整數。")
                continue

            hold_minutes_input = input("請輸入持有時間（分鐘，輸入 'F' 代表持有到13:30強制出場）：")
            if hold_minutes_input.upper() == 'F':
                hold_minutes = None
            else:
                try:
                    hold_minutes = int(hold_minutes_input)
                except ValueError:
                    print("持有時間必須是整數或 'F'。")
                    continue

            disposition_stocks = load_disposition_stocks()
            symbols_to_analyze = matrix_dict_analysis[group_name]
            symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]
            if len(symbols_to_analyze) == 0:
                print(f"{group_name} 中沒有可供分析的股票。")
                continue

            daily_kline_data, intraday_kline_data = load_kline_data()

            stock_data_collection = initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data)
            if not stock_data_collection:
                print("無法獲取有效的一分 K 資料，無法進行分析")
                continue

            total_profit, avg_profit_rate = process_group_data(stock_data_collection, wait_minutes, hold_minutes, matrix_dict_analysis, verbose=True)

            print(f"\n模擬交易完成，總利潤：{int(total_profit) if total_profit is not None else 0} 元，平均報酬率：{avg_profit_rate if avg_profit_rate is not None else 0:.2f}%\n")

        elif choice == '2':
            try:
                wait_minutes = int(input("請輸入等待時間（分鐘）："))
            except ValueError:
                print("等待時間必須是整數。")
                continue

            hold_minutes_input = input("請輸入持有時間（分鐘，輸入 'F' 代表持有到13:30強制出場）：")
            if hold_minutes_input.upper() == 'F':
                hold_minutes = None
            else:
                try:
                    hold_minutes = int(hold_minutes_input)
                except ValueError:
                    print("持有時間必須是整數或 'F'。")
                    continue

            day_total_profit = 0
            day_avg_profit_rates = []

            for group_name in matrix_dict_analysis.keys():
                print(f"\n正在分析族群：{group_name}")

                disposition_stocks = load_disposition_stocks()
                symbols_to_analyze = matrix_dict_analysis[group_name]
                symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]
                if len(symbols_to_analyze) == 0:
                    print(f"{group_name} 中沒有可供分析的股票。")
                    continue

                daily_kline_data, intraday_kline_data = load_kline_data()

                stock_data_collection = initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data)
                if not stock_data_collection:
                    print(f"無法獲取 {group_name} 的有效一分 K 資料，跳過。")
                    continue

                total_profit, avg_profit_rate = process_group_data(stock_data_collection, wait_minutes, hold_minutes, matrix_dict_analysis, verbose=True)

                if total_profit is not None and avg_profit_rate is not None:
                    day_total_profit += total_profit
                    day_avg_profit_rates.append(avg_profit_rate)
                else:
                    pass

            if day_avg_profit_rates:
                day_avg_profit_rate = sum(day_avg_profit_rates) / len(day_avg_profit_rates)
            else:
                day_avg_profit_rate = 0.0

            if day_total_profit > 0:
                print(f"{RED}=" * 50)
                print(f"{RED}\n當日總利潤：{int(day_total_profit)} 元{RESET}")
                print(f"{RED}當日報酬率：{day_avg_profit_rate:.2f}%\n{RESET}")
                print(f"{RED}=" * 50)
            elif day_total_profit < 0:
                print(f"{GREEN}=" * 50)
                print(f"{GREEN}\n當日總利潤：{int(day_total_profit)} 元{RESET}")
                print(f"{GREEN}當日報酬率：{day_avg_profit_rate:.2f}%\n{RESET}")
                print(f"{GREEN}=" * 50)
            else:
                print("=" * 50)
                print(f"\n當日總利潤：{int(day_total_profit)} 元")
                print(f"當日報酬率：{day_avg_profit_rate:.2f}%\n")
                print("=" * 50)

        elif choice == '0':
            break
        else:
            print("無效的選項，請重新輸入。")

# --------------- 回測程式：極大化利潤模式 ---------------
def maximize_profit_analysis():
    print("進入極大化利潤模式...")
    
    matrix_dict_analysis = load_matrix_dict_analysis()
    if not matrix_dict_analysis:
        print("沒有族群資料，請先管理族群。")
        return

    group_name = input("請輸入要分析的族群名稱：")
    
    if group_name not in matrix_dict_analysis:
        print("沒有此族群資料")
        return
    wait_minutes_start = int(input("請輸入等待時間起始值（分鐘）："))
    wait_minutes_end = int(input("請輸入等待時間結束值（分鐘）："))
    hold_minutes_start = int(input("請輸入持有時間起始值（分鐘，輸入0代表F）："))
    hold_minutes_end = int(input("請輸入持有時間結束值（分鐘，輸入0代表F）："))

    wait_minutes_range = range(wait_minutes_start, wait_minutes_end + 1)
    hold_minutes_range = range(hold_minutes_start, hold_minutes_end + 1)

    disposition_stocks = load_disposition_stocks()
    symbols_to_analyze = matrix_dict_analysis[group_name]
    symbols_to_analyze = [symbol for symbol in symbols_to_analyze if symbol not in disposition_stocks]
    if len(symbols_to_analyze) == 0:
        print(f"{group_name} 中沒有可供分析的股票。")
        return

    daily_kline_data, intraday_kline_data = load_kline_data()

    stock_data_collection = initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data)
    if not stock_data_collection:
        print("無法獲取有效的一分 K 資料，無法進行分析")
        return

    results_df = pd.DataFrame(columns=['等待時間', '持有時間', '總利潤', '平均報酬率'])
    results_df = results_df.astype({
        '等待時間': 'int',
        '持有時間': 'object',
        '總利潤': 'float',
        '平均報酬率': 'float'
    })

    for wait_minutes in wait_minutes_range:
        for hold_minutes in hold_minutes_range:
            hold_minutes_value = None if hold_minutes == 0 else hold_minutes
            print(f"正在分析：等待時間 {wait_minutes} 分鐘、持有時間 {'F' if hold_minutes_value is None else hold_minutes_value} 分鐘")
            
            total_profit, avg_profit_rate = process_group_data(
                stock_data_collection, wait_minutes, hold_minutes_value, matrix_dict_analysis, verbose=False)
            
            if total_profit is None:
                total_profit = 0.0
            if avg_profit_rate is None:
                avg_profit_rate = 0.0
            
            new_row = pd.DataFrame([{
                '等待時間': wait_minutes,
                '持有時間': 'F' if hold_minutes_value is None else hold_minutes_value,
                '總利潤': float(total_profit),
                '平均報酬率': float(avg_profit_rate)
            }])
            results_df = pd.concat([results_df, new_row], ignore_index=True)

    if results_df.empty:
        print("模擬結果為空，無法進行後續分析。")
        return

    max_profit = results_df['總利潤'].max()
    min_profit = results_df['總利潤'].min()
    best_combination = results_df.loc[results_df['總利潤'].idxmax()]

    print("\n利潤最大的組合：")
    print(f"等待時間：{best_combination['等待時間']} 分鐘，持有時間：{best_combination['持有時間']} 分鐘，總利潤：{int(best_combination['總利潤'])} 元，平均報酬率：{best_combination['平均報酬率']:.2f}%\n")

    pivot_df = results_df.pivot(index='等待時間', columns='持有時間', values='總利潤')

    formatted_pivot_df = pivot_df.copy()
    for col in formatted_pivot_df.columns:
        if col != '等待時間':
            formatted_pivot_df[col] = formatted_pivot_df[col].apply(lambda x: f"{int(x):,}" if pd.notnull(x) else "")

    formatted_pivot_df_reset = formatted_pivot_df.reset_index()

    print("模擬結果：")
    print(tabulate(formatted_pivot_df_reset, headers='keys', tablefmt='psql', showindex=False))

    try:
        with pd.ExcelWriter('模擬結果.xlsx', engine='openpyxl') as writer:
            pivot_df.to_excel(writer, sheet_name='模擬結果', index=True)
            workbook = writer.book
            worksheet = writer.sheets['模擬結果']
            
            max_profit = pivot_df.max().max()
            min_profit = pivot_df.min().min()

            max_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            min_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

            for row in worksheet.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    if cell.value == max_profit:
                        cell.fill = max_fill
                    elif cell.value == min_profit:
                        cell.fill = min_fill
        print("\n模擬結果已成功寫入 '模擬結果.xlsx'。")
    except Exception as e:
        print(f"\n寫入 Excel 時發生錯誤：{e}")




# ------------------ 交易程式 ------------------
def trading_menu_list():
    print('\n' + '=' * 50)
    print("\n請選擇功能：")
    print("1. 開始交易、2. 登入帳戶、0. 返回主選單")
    print('\n' + '=' * 50)
    back_choice = input("請選擇功能：")
    if back_choice == '1':
        start_trading()
    elif back_choice == '2':
        login()
    elif back_choice == '0':
        main_menu()
    else:
        print("無效的選項，請重新輸入")

# ------------------ 交易程式：開始交易 ------------------
def start_trading(mode='full', wait_minutes=None, hold_minutes=None):
    """
    mode:
        'full' – 第一次執行：正常詢問等待/持有分鐘。
        'post' – 盤後遞迴呼叫：沿用上一輪 wait_minutes / hold_minutes，不再詢問。
    """
    client, api_key = init_fugle_client()

    # ===== 處置股過濾=====
    matrix_dict_analysis = load_matrix_dict_analysis()
    fetch_disposition_stocks(client, matrix_dict_analysis)   # ① 先更新 Disposition.json
    disposition_stocks = load_disposition_stocks()           # ② 讀最新處置股
    purge_disposition_from_nb(disposition_stocks)           # ③ 刪 nb_matrix_dict 中的處置股
    # ====================

    symbols_to_analyze = load_symbols_to_analyze()
    stop_trading = False
    max_symbols_to_fetch = 20

    group_symbols = load_group_symbols()
    if not group_symbols:
        print("沒有加載到任何族群資料，請確認 nb_matrix_dict.json 的存在與內容。")
        return
    consolidated_symbols = group_symbols.get('consolidated_symbols', {})
    if not consolidated_symbols:
        print("沒有找到 'consolidated_symbols'，請確認資料結構。")
        return
    group_positions = {group: False for group in consolidated_symbols.keys()}

    # 時間判斷
    now = datetime.now()
    now_str = now.strftime('%Y-%m-%d %H:%M:%S')
    pre_market_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    market_start     = now.replace(hour=9, minute=1, second=0, microsecond=0)
    market_end       = now.replace(hour=13, minute=30, second=0, microsecond=0)
    post_switch      = now.replace(hour=13, minute=32, second=0, microsecond=0)
    market_exit      = now.replace(hour=13, minute=26, second=0, microsecond=0)

    # 先分支：盤前、盤中、轉盤後過渡、盤後
    if pre_market_start <= now < market_start:
        print(f"目前為 {now_str}，盤前時間，只更新日K線資料。")
        
        # ---------- 取得 / 比對日 K（盤前） ----------
        existing_auto_daily_data = {}
        if os.path.exists('auto_daily.json'):
            with open('auto_daily.json', 'r', encoding='utf-8') as f:
                try:
                    existing_auto_daily_data = json.load(f)
                except json.JSONDecodeError:
                    existing_auto_daily_data = {}
        else:
            print("auto_daily.json 不存在，將建立新的。")

        print("開始取得日K線數據並與現有資料比對...")
        auto_daily_data = {}
        data_is_same = True
        initial_api_count = 0
        symbols_fetched = 0

        for symbol in symbols_to_analyze[:max_symbols_to_fetch]:
            if initial_api_count >= 55:
                print("已達到55次API請求，休息1分鐘...")
                time_module.sleep(60)
                initial_api_count = 0
            daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
            initial_api_count += 1
            if daily_kline_df.empty:
                print(f"無法取得 {symbol} 的日K數據，跳過。")
                continue
            daily_kline_data = daily_kline_df.to_dict(orient='records')
            auto_daily_data[symbol] = daily_kline_data
            existing_data = existing_auto_daily_data.get(symbol)
            if existing_data != daily_kline_data:
                data_is_same = False
                print(f"{symbol} 的日K數據與現有資料不同，將更新資料。")
                existing_auto_daily_data[symbol] = daily_kline_data
            else:
                print(f"{symbol} 的日K數據與現有資料相同，跳過更新。")
            symbols_fetched += 1

        if not data_is_same:
            remaining_symbols = symbols_to_analyze[max_symbols_to_fetch:]
            print(f"發現前 {max_symbols_to_fetch} 支股票的日K數據有更新，開始取得剩餘股票的日K數據並更新。")
            for symbol in remaining_symbols:
                if initial_api_count >= 55:
                    print("已達到55次API請求，休息1分鐘...")
                    time_module.sleep(60)
                    initial_api_count = 0
                daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
                initial_api_count += 1
                if daily_kline_df.empty:
                    print(f"無法取得 {symbol} 的日K數據，跳過。")
                    continue
                daily_kline_data = daily_kline_df.to_dict(orient='records')
                auto_daily_data[symbol] = daily_kline_data
                existing_data = existing_auto_daily_data.get(symbol)
                if existing_data != daily_kline_data:
                    print(f"{symbol} 的日K數據與現有資料不同，將更新資料。")
                    existing_auto_daily_data[symbol] = daily_kline_data
                else:
                    print(f"{symbol} 的日K數據與現有資料相同，跳過更新。")

        if symbols_fetched < max_symbols_to_fetch:
            print(f"注意：僅取得了 {symbols_fetched} 支股票的日K數據。")

        with open('auto_daily.json', 'w', encoding='utf-8') as f:
            json.dump(existing_auto_daily_data, f, ensure_ascii=False, indent=4)

        print(f"{YELLOW}已更新 auto_daily.json。{RESET}")
        print(f"{YELLOW}盤前更新完成。{RESET}")

        # ✅ 這邊新增：等待到 09:00 自動切入盤中交易模式
        now = datetime.now()
        wait_seconds = (market_start - now).total_seconds()
        if wait_seconds > 0:
            print(f"等待 {wait_seconds/60:.1f} 分鐘直到開盤開始盤中交易...")
            time_module.sleep(wait_seconds)

        # ✅ 等待完後自動更新現在時間，進入盤中流程
        now = datetime.now()


    elif market_start <= now <= market_end:
        print(f"目前為 {now_str}，盤中交易時間。")
        # ---------- 1. 第一次執行詢問使用者 ----------
        if mode == 'full':
            try:
                wait_minutes = int(input("請輸入等待時間（分鐘）："))
            except ValueError:
                print("等待時間必須是整數。")
                return
            hold_minutes_input = input("請輸入持有時間（分鐘，輸入 'F' 代表持有到13:30強制出場）：")
            if hold_minutes_input.upper() == 'F':
                hold_minutes = None
            else:
                try:
                    hold_minutes = int(hold_minutes_input)
                except ValueError:
                    print("持有時間必須是整數或 'F'。")
                    return
        else:
            assert wait_minutes is not None

        # ---------- 2. 取得 / 比對日 K（盤中也需要日K） ----------
        existing_auto_daily_data = {}
        if os.path.exists('auto_daily.json'):
            with open('auto_daily.json', 'r', encoding='utf-8') as f:
                try:
                    existing_auto_daily_data = json.load(f)
                except json.JSONDecodeError:
                    existing_auto_daily_data = {}
        print("開始取得日K線數據並與現有資料比對...")
        auto_daily_data = {}
        data_is_same = True
        initial_api_count = 0
        symbols_fetched = 0
        for symbol in symbols_to_analyze[:max_symbols_to_fetch]:
            if initial_api_count >= 55:
                print("已達到55次API請求，休息1分鐘...")
                time_module.sleep(60)
                initial_api_count = 0
            daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
            initial_api_count += 1
            if daily_kline_df.empty:
                print(f"無法取得 {symbol} 的日K數據，跳過。")
                continue
            daily_kline_data = daily_kline_df.to_dict(orient='records')
            auto_daily_data[symbol] = daily_kline_data
            existing_data = existing_auto_daily_data.get(symbol)
            if existing_data != daily_kline_data:
                data_is_same = False
                print(f"{symbol} 的日K數據與現有資料不同，將更新資料。")
                existing_auto_daily_data[symbol] = daily_kline_data
            else:
                print(f"{symbol} 的日K數據與現有資料相同，跳過更新。")
            symbols_fetched += 1
        if not data_is_same:
            remaining_symbols = symbols_to_analyze[max_symbols_to_fetch:]
            print(f"發現前 {max_symbols_to_fetch} 支股票的日K數據有更新，開始取得剩餘股票的日K數據並更新。")
            for symbol in remaining_symbols:
                if initial_api_count >= 55:
                    print("已達到55次API請求，休息1分鐘...")
                    time_module.sleep(60)
                    initial_api_count = 0
                daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
                initial_api_count += 1
                if daily_kline_df.empty:
                    print(f"無法取得 {symbol} 的日K數據，跳過。")
                    continue
                daily_kline_data = daily_kline_df.to_dict(orient='records')
                auto_daily_data[symbol] = daily_kline_data
                existing_data = existing_auto_daily_data.get(symbol)
                if existing_data != daily_kline_data:
                    print(f"{symbol} 的日K數據與現有資料不同，將更新資料。")
                    existing_auto_daily_data[symbol] = daily_kline_data
                else:
                    print(f"{symbol} 的日K數據與現有資料相同，跳過更新。")
        if symbols_fetched < max_symbols_to_fetch:
            print(f"注意：僅取得了 {symbols_fetched} 支股票的日K數據。")
        with open('auto_daily.json', 'w', encoding='utf-8') as f:
            json.dump(existing_auto_daily_data, f, ensure_ascii=False, indent=4)
        print("已更新 auto_daily.json。")

        # ---------- 3. 補齊一分K（盤中每次都要對當日做初次補齊） ----------
        fetch_time = datetime.now() - timedelta(minutes=1)
        trading_day = fetch_time.strftime('%Y-%m-%d')
        '''
        print(f"日期樣本：{trading_day}")
        '''
        
        yesterday_close_prices = {}
        # （此處保持「讀昨收」邏輯不變）
        for symbol in symbols_to_analyze:
            daily_data = existing_auto_daily_data.get(symbol, [])
            if not daily_data:
                daily_kline_df = fetch_daily_kline_data(client, symbol, days=5)
                if not daily_kline_df.empty:
                    daily_kline_data = daily_kline_df.to_dict(orient='records')
                    existing_auto_daily_data[symbol] = daily_kline_data
                    with open('auto_daily.json', 'w', encoding='utf-8') as f:
                        json.dump(existing_auto_daily_data, f, ensure_ascii=False, indent=4)
                if len(existing_auto_daily_data[symbol]) > 1:
                    now2 = datetime.now()
                    weekday = now2.weekday()
                    if 0 <= weekday <= 4 and 8 <= now2.hour < 15:
                        yesterday_close = existing_auto_daily_data[symbol][0].get('close', 0)
                    else:
                        yesterday_close = existing_auto_daily_data[symbol][1].get('close', 0)
                else:
                    yesterday_close = 0
                yesterday_close_prices[symbol] = yesterday_close
            else:
                sorted_daily_data = sorted(daily_data, key=lambda x: x['date'], reverse=True)
                if len(sorted_daily_data) > 1:
                    now2 = datetime.now()
                    weekday = now2.weekday()
                    if 0 <= weekday <= 4 and 8 <= now2.hour < 15:
                        yesterday_close = sorted_daily_data[0].get('close', 0)
                    else:
                        yesterday_close = sorted_daily_data[1].get('close', 0)
                else:
                    yesterday_close = sorted_daily_data[0].get('close', 0)
                yesterday_close_prices[symbol] = yesterday_close

        # 一分K初次補齊

        # 測試訊息
        t_fetch_hist = time_module.perf_counter()
        print("🔁 [歷史] 開始補齊一分K資料...")
        
        market_real_end       = now.replace(hour=13, minute=30, second=0, microsecond=0)

        if now < market_real_end :
            full_intraday_end = (now - timedelta(minutes=1)).strftime('%H:%M')
        else:
            full_intraday_end = "13:30"


        print(f"{YELLOW}開始補齊今日 09:00 到 {full_intraday_end} 的一分K數據...{RESET}")

        auto_intraday_data = {}
        initial_api_count = 0
        with ThreadPoolExecutor(max_workers=200) as executor:
            future_to_symbol = {}
            for symbol in symbols_to_analyze:
                if initial_api_count >= 200:
                    time_module.sleep(60)
                    initial_api_count = 0
                yc = yesterday_close_prices.get(symbol, 0)
                if yc == 0:
                    continue
                future = executor.submit(
                    fetch_intraday_data,
                    client=client,
                    symbol=symbol,
                    trading_day=trading_day,
                    yesterday_close_price=yc,
                    start_time="09:00",
                    end_time=full_intraday_end
                )
                future_to_symbol[future] = symbol
                initial_api_count += 1
            for future in as_completed(future_to_symbol):
                symbol = future_to_symbol[future]
                df = future.result()
                if df.empty:
                    continue
                updated_records = []
                records = df.to_dict(orient='records')
                for i, candle in enumerate(records):
                    past_candles = records[:i]
                    updated_candle = calculate_2min_pct_increase_and_highest(candle, past_candles)
                    updated_records.append(updated_candle)
                df = pd.DataFrame(updated_records)
                auto_intraday_data[symbol] = df.to_dict(orient='records')

        # 測試訊息
        print(f"✅ [歷史] 補齊完成，耗時：{time_module.perf_counter() - t_fetch_hist:.2f} 秒")
        t_save_json = time_module.perf_counter()

        save_auto_intraday_data(auto_intraday_data)

        initialize_triggered_limit_up(auto_intraday_data)

        # 測試訊息
        print(f"📝 [寫檔] 寫入 auto_intraday.json 完成，耗時：{time_module.perf_counter() - t_save_json:.2f} 秒")
        '''
        print("已更新 auto_intraday.json。")
        '''
        # ---------- 4. 盤中主迴圈 ----------
        print("開始盤中交易監控，輸入 'Q' 返回主選單： ", end='', flush=True)

        # 啟動非阻塞 Q 鍵監聽與選單觸發
        threading.Thread(target=monitor_quit_key, daemon=True).start()
        threading.Thread(target=check_quit_flag_loop, daemon=True).start()

        # 初始化盤中狀態
        has_exited = False
        current_position = None
        hold_time = 0
        message_log = []
        already_entered_stocks = []
        stop_loss_triggered = False
        final_check_active = False
        final_check_count = 0
        in_waiting_period = False
        waiting_time = 0
        leader = None
        tracking_stocks = set()
        previous_rise_values = {}
        leader_peak_rise = None
        leader_rise_before_decline = None
        first_condition_one_time = None
        can_trade = True

        exit_live_done = False
        restart_to_post = False

        while not stop_trading:
            now_loop = datetime.now()

            if now_loop >= market_exit and not exit_live_done:
                print(f"🔍 13:26 觸發：檢查觸價委託單，目前尚有 {len(to.conditions)} 檔股票在觸價委託中。")
                exit_trade_live()
                exit_live_done = True

            if market_end < now_loop < post_switch:
                print(f"\n目前為 {now_loop.strftime('%Y-%m-%d %H:%M:%S')}，盤後過渡期，等待切盤後流程…")
                time_module.sleep((post_switch - now_loop).total_seconds())
                continue

            if now_loop >= post_switch:
                print("\n收盤後 +1 分鐘，切換到盤後流程…")
                restart_to_post = True
                break

            if market_start <= now_loop <= market_end:
                now_sec = datetime.now().second
                time_module.sleep(60 - now_sec)

                fetch_time = datetime.now() - timedelta(minutes=1)
                trading_day = fetch_time.strftime('%Y-%m-%d')
                fetch_time_str = fetch_time.strftime('%H:%M')
                if fetch_time.time() > market_end.time():
                    fetch_time_str = "13:30"
                
                # 測試訊息
                t_fetch_realtime = time_module.perf_counter()
                print(f"{YELLOW}⏱ [即時] 開始取得 {fetch_time_str} 的一分K資料...{RESET}")

                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                print("\n" + "=" * 50)
                print(f"\n{timestamp} 市場開盤中，取得 {fetch_time_str} 分的即時一分K數據。")
                print(f"正在取得一分K數據從 {fetch_time_str} 到 {fetch_time_str}...")

                updated_intraday_data = {}
                with ThreadPoolExecutor(max_workers=200) as executor:
                    future_to_symbol = {}
                    for symbol in symbols_to_analyze:
                        yc = yesterday_close_prices.get(symbol, 0)
                        if yc == 0:
                            continue
                        fut = executor.submit(
                            fetch_realtime_intraday_data,
                            client=client,
                            symbol=symbol,
                            trading_day=trading_day,
                            yesterday_close_price=yc,
                            start_time=fetch_time_str,
                            end_time=fetch_time_str
                        )
                        future_to_symbol[fut] = symbol
                    for fut in as_completed(future_to_symbol):
                        sym = future_to_symbol[fut]
                        df = fut.result()
                        if df.empty:
                            continue
                        candle = df.to_dict(orient='records')[0]
                        candle = calculate_2min_pct_increase_and_highest(candle, auto_intraday_data.get(sym, []))
                        if '漲停價' in candle:
                            candle['漲停價'] = truncate_to_two_decimals(candle['漲停價'])
                        updated_intraday_data.setdefault(sym, []).append(candle)

                for sym, lst in updated_intraday_data.items():
                    auto_intraday_data.setdefault(sym, []).extend(lst)
                    auto_intraday_data[sym] = auto_intraday_data[sym][-1000:]

                # 測試訊息
                print(f"✅ [即時] 一分K取得完成，耗時：{time_module.perf_counter() - t_fetch_realtime:.2f} 秒")
                t_save_json = time_module.perf_counter()

                save_auto_intraday_data(auto_intraday_data)

                # 測試訊息
                print(f"📝 [寫檔] 寫入 auto_intraday.json 完成，耗時：{time_module.perf_counter() - t_save_json:.2f} 秒")
                print("=" * 50)

                process_live_trading_logic(
                    symbols_to_analyze,
                    fetch_time_str,
                    wait_minutes,
                    hold_minutes,
                    message_log,
                    False,
                    has_exited,
                    current_position,
                    hold_time,
                    already_entered_stocks,
                    stop_loss_triggered,
                    final_check_active,
                    final_check_count,
                    in_waiting_period,
                    waiting_time,
                    leader,
                    tracking_stocks,
                    previous_rise_values,
                    leader_peak_rise,
                    leader_rise_before_decline,
                    first_condition_one_time,
                    can_trade,
                    group_positions
                )

        # 若有切換到盤後
        if restart_to_post:
            start_trading(mode='post', wait_minutes=wait_minutes, hold_minutes=hold_minutes)
            return

        print("已停止交易，返回主選單")

    else:  # now >= post_switch
        print(f"目前為 {now_str}，盤後時間，不需要更新任何數據，返回主選單。")
        return

# ------------------ 交易程式：登入帳戶 ------------------
def login():
    file_path = "shioaji_logic.py"  # 要更新的檔案路徑

    print('\n' + '=' * 50 + '\n')
    print("當前 api key 為：" + shioaji_logic.TEST_API_KEY)
    print("當前憑證路徑為：" + shioaji_logic.CA_CERT_PATH)
    print("當前憑證密碼為：" + shioaji_logic.CA_PASSWORD)
    print('\n' + '=' * 50)
    print("1. 修改 api key、2. 修改 api secret、3. 修改憑證路徑、4. 修改憑證密碼")
    api_setting = input("請選擇功能：")
    if api_setting == "1":
        new_api_key = input("請輸入新的 api key：")
        update_variable(file_path, "TEST_API_KEY", new_api_key)
    elif api_setting == "2":
        new_api_secret = input("請輸入新的 api secret：")
        update_variable(file_path, "TEST_API_SECRET", new_api_secret)
    elif api_setting == "3":
        new_ca_path = input("請輸入新的憑證路徑：")
        update_variable(file_path, "CA_CERT_PATH", new_ca_path, is_raw=True)
    elif api_setting == "4":
        new_ca_password = input("請輸入新的憑證密碼：")
        update_variable(file_path, "CA_PASSWORD", new_ca_password)
    else:
        print("請輸入合法字元...")
        login()

def update_variable(file_path, var_name, new_value, is_raw=False):
    """
    更新指定檔案中以 var_name 開頭的變數的值。
    若 is_raw 為 True，則會以 raw 字串格式儲存（例如 CA_CERT_PATH）
    """
    lines = []
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            # 如果該行以變數名稱開頭，則替換該行
            if line.lstrip().startswith(var_name + " ="):
                if is_raw:
                    new_line = f'{var_name} = r"{new_value}"\n'
                else:
                    new_line = f'{var_name} = "{new_value}"\n'
                lines.append(new_line)
            else:
                lines.append(line)
    with open(file_path, "w", encoding="utf-8") as f:
        f.writelines(lines)
    print(f"{var_name} 已更新為: {new_value}")
    importlib.reload(shioaji_logic)

#登入
api = sj.Shioaji(simulation=True)
accounts = api.login(api_key = shioaji_logic.TEST_API_KEY, secret_key = shioaji_logic.TEST_API_SECRET)
api.activate_ca(
    ca_path=shioaji_logic.CA_CERT_PATH,
    ca_passwd=shioaji_logic.CA_PASSWORD
)
'''
print("ca_path:", shioaji_logic.CA_CERT_PATH)
print("ca_password:", shioaji_logic.CA_PASSWORD)
'''




# ------------------ 管理族群 ------------------
def manage_groups():
    current_page = 0
    page_size = 5
    groups = load_matrix_dict_analysis()
    total_pages = (len(groups) + page_size - 1) // page_size
    total_page = 1

    def display_page(page):
        load_twse_name_map()                     # ← 確保已載入對照表
        start = page * page_size
        end   = start + page_size
        if total_pages == 0:
            print("=" * 50)
            print(f"族群及個股列表 - 第 {page + 1} 頁 / 共 {total_page} 頁")
            print("=" * 50)
        else:
            print("=" * 50)
            print(f"族群及個股列表 - 第 {page + 1} 頁 / 共 {total_pages} 頁")
            print("=" * 50)
        for idx, (group, stocks) in enumerate(list(groups.items())[start:end], start=1):
            print(f"族群: {group}")
            if stocks:
                for s_idx, code in enumerate(stocks, start=1):
                    cname = get_stock_name(code)
                    print(f"  {str(s_idx).rjust(2)}. {code:<6} {cname}")
            else:
                print("  (此族群目前沒有個股)")
            print("-" * 50)
        print("=" * 50)
        if current_page == total_pages - 1:
            print("已顯示所有族群及個股。")
        print("=" * 50)

    while True:
        display_page(current_page)
        print("\nP：上一頁、Q：下一頁、1：新增族群/個股；、2：刪除族群/個股、0：返回主選單")
        choice = input("請選擇操作: ").strip()

        if choice == "1":
            add_group_or_stock(groups)
        elif choice == "2":
            delete_group_or_stock(groups)
        elif choice.upper() == "P":
            if current_page > 0:
                current_page -= 1
            else:
                print("已經是第一頁！")
        elif choice.upper() == "Q":
            if current_page < total_pages - 1:
                current_page += 1
            else:
                print("已經是最後一頁！")
        elif choice == "0":
            save_matrix_dict(groups)
            break
        else:
            print("無效選項，請重新選擇。")

#------------------ 管理族群：新增族群/個股 ------------------
def add_group_or_stock(groups):
    print("\n==============================")
    print("1：新增族群、2：新增族群中的個股、0：返回選單")
    print("\n==============================")
    choice = input("請選擇操作: ").strip()

    if choice == "1":
        new_group = input("輸入新族群名稱: ").strip()
        if not new_group:
            print("族群名稱不能為空。")
            add_group_or_stock(groups)
        if new_group in groups:
            print(f"族群 '{new_group}' 已存在。")
        else:
            groups[new_group] = []
            print(f"族群 '{new_group}' 新增成功。")
    
    elif choice == "2":
        group_name = input("輸入要新增個股的族群名稱: ").strip()
        if not group_name:
            print("族群名稱不能為空。")
            add_group_or_stock(groups)
        if group_name in groups:
            current_stocks = groups[group_name]
            print(f"\n==============================")
            print(f"族群 '{group_name}' 中目前的個股:")
            if current_stocks:
                for idx, stock in enumerate(current_stocks, start=1):
                    print(f"  {str(idx).rjust(2)}. {stock}")
            else:
                print("  無")
            print("==============================\n")
            
            print(f"開始新增個股到族群 '{group_name}'。")
            print("請輸入個股代號，輸入 'Q' 以退出新增模式。")
            
            while True:
                new_stock = input("輸入個股代號 (或 'Q' 退出): ").strip()
                if new_stock.upper() == "Q":
                    print("退出新增個股模式。")
                    break
                elif not new_stock:
                    print("輸入無效，請重新輸入。")
                    continue
                elif new_stock in groups[group_name]:
                    print(f"個股 '{new_stock} {get_stock_name(new_stock)}' 已存在於族群 '{group_name}' 中。")
                else:
                    groups[group_name].append(new_stock)
                    print(f"個股 '{new_stock} {get_stock_name(new_stock)}' 已新增至族群 '{group_name}'。")
        else:
            print(f"族群 '{group_name}' 不存在。")
    
    elif choice == "0":
        print("返回主選單。")
        manage_groups()

    else:
        print("無效的選項，請重新選擇。")

#------------------ 管理族群：刪除族群/個股 ------------------
def delete_group_or_stock(groups):
    print("\n==============================")
    print("1：刪除族群、2：刪除族群中的個股、0：返回選單")
    print("\n==============================")
    choice = input("請選擇操作: ").strip()

    if choice == "1":
        group_name = input("輸入要刪除的族群名稱: ").strip()
        if not group_name:
            print("族群名稱不能為空。")
            delete_group_or_stock(groups)
        if group_name in groups:
            confirm = input(f"確定要刪除族群 '{group_name}' 嗎？ (Y/N): ").strip().upper()
            if confirm == "Y":
                del groups[group_name]
                print(f"族群 '{group_name}' 已刪除。")
            else:
                print("取消刪除。")
        else:
            print(f"族群 '{group_name}' 不存在。")

    elif choice == "2":
        group_name = input("輸入要刪除個股的族群名稱: ").strip()
        if not group_name:
            print("族群名稱不能為空。")
            delete_group_or_stock(groups)
        if group_name in groups:
            current_stocks = groups[group_name]
            print(f"\n==============================")
            print(f"族群 '{group_name}' 中目前的個股:")
            if current_stocks:
                for idx, stock in enumerate(current_stocks, start=1):
                    print(f"  {str(idx).rjust(2)}. {stock}")
            else:
                print("  無")
            print("==============================\n")

            if not current_stocks:
                print(f"族群 '{group_name}' 中目前沒有任何個股。")
                delete_group_or_stock(groups)

            print(f"開始刪除個股從族群 '{group_name}'。")
            print("請輸入要刪除的個股代號，輸入 'Q' 以退出刪除模式。")

            while True:
                stock_name = input("輸入個股代號 (或 'Q' 退出): ").strip()
                if stock_name.upper() == "Q":
                    print("退出刪除個股模式。")
                    break
                elif not stock_name:
                    print("輸入無效，請重新輸入。")
                    continue
                elif stock_name not in groups[group_name]:
                    print(f"個股 '{stock_name}' 不存在於族群 '{group_name}' 中。")
                else:
                    confirm = input(f"確定要刪除個股 '{stock_name} {get_stock_name(stock_name)} '嗎？ (Y/N): ").strip().upper()
                    if confirm == "Y":
                        groups[group_name].remove(stock_name)
                        print(f"個股 '{stock_name}' 已從族群 '{group_name}' 中刪除。")
                        if not groups[group_name]:
                            print(f"族群 '{group_name}' 現在已經沒有任何個股。")
                    else:
                        print("取消刪除。")
        else:
            print(f"族群 '{group_name}' 不存在。")

    elif choice == "0":
        print("返回主選單。")
        manage_groups()

    else:
        print("無效的選項，請重新選擇。")




# ------------------ 設定選單 ------------------
def settings_menu():
    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    global allow_reentry_after_stop_loss
    while True:
        print("\n設定選單：")
        print(f"1. 設定每檔股票投入資本額（目前為 {capital_per_stock} 萬元）")
        print(f"2. 手續費設定，目前為 {transaction_fee}%")
        print(f"3. 手續費折數設定，目前為 {transaction_discount}%")
        print(f"4. 證交稅設定，目前為 {trading_tax}%")
        print("5. 價差停損設定")
        print("6. 停損再進場設定")
        print("0. 返回主選單")
        choice = input("請輸入選項：")
        if choice == "1":
            set_capital_per_stock()
        elif choice == "2":
            transaction_fee = float(input("請輸入手續費（%）："))
            save_settings()
        elif choice == "3":
            transaction_discount = float(input("請輸入手續費折數（%）："))
            save_settings()
        elif choice == "4":
            trading_tax = float(input("請輸入證交稅（%）："))
            save_settings()
        elif choice == "5":
            price_gap_stop_loss_menu()
        elif choice == "6":
            stop_loss_reentry_menu()
        elif choice == "0":
            main_menu()
        else:
            print("無效的選項，請重新輸入")

# ------------------ 設定選單：設定每檔資本額 ------------------
def set_capital_per_stock():
    global capital_per_stock
    capital_per_stock = int(input("請輸入每檔投入資本額（萬元）："))
    print(f"每檔投入資本額已設定為：{capital_per_stock} 萬元")
    save_settings()

# ------------------ 設定選單：價差停損設定 ------------------
def price_gap_stop_loss_menu():
    global below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    while True:
        print(f"1. 50元以下股票停損價差，目前為 {below_50} 元")
        print(f"2. 50~100元股票停損價差，目前為 {price_gap_50_to_100} 元")
        print(f"3. 100~500元股票停損價差，目前為 {price_gap_100_to_500} 元")
        print(f"4. 500~1000元股票停損價差，目前為 {price_gap_500_to_1000} 元")
        print(f"5. 1000元以上股票停損價差，目前為 {price_gap_above_1000} 元")
        print("6. 返回上一頁")
        choice = input("請選擇要設定的項目：")
        if choice == "1":
            below_50 = float(input("請輸入50元以下股票的停損價差："))
        elif choice == "2":
            price_gap_50_to_100 = float(input("請輸入50~100元股票的停損價差："))
        elif choice == "3":
            price_gap_100_to_500 = float(input("請輸入100~500元股票的停損價差："))
        elif choice == "4":
            price_gap_500_to_1000 = float(input("請輸入500~1000元股票的停損價差："))
        elif choice == "5":
            price_gap_above_1000 = float(input("請輸入1000元以上股票的停損價差："))
        elif choice == "6":
            break
        else:
            print("無效選擇，請重試。")
        save_settings()

# ------------------ 設定選單：停損再進場設定 ------------------
def stop_loss_reentry_menu():
    global allow_reentry_after_stop_loss
    while True:
        status = "開啟" if allow_reentry_after_stop_loss else "關閉"
        print(f"\n目前為({status}停損後進場)")
        print("1.開啟停損後進場")
        print("2.關閉停損後進場")
        print("3.返回上一頁")
        choice = input("請輸入選項：")
        if choice == '1':
            allow_reentry_after_stop_loss = True
            print("已開啟停損後進場功能")
            save_settings()
        elif choice == '2':
            allow_reentry_after_stop_loss = False
            print("已關閉停損後進場功能")
            save_settings()
        elif choice == '3':
            settings_menu()
        else:
            print("無效的選項，請重新輸入")




# ------------------ 更新K線數據 ------------------
def update_kline_data_menu():
    while True:
        print("\n更新K線數據選單：")
        print("1. 更新數據")
        print("2. 查看數據")
        print("0. 返回主選單")
        choice = input("請輸入選項：")
        if choice == '1':
            update_kline_data()
        elif choice == '2':
            symbol_to_group = load_symbol_to_group('./matrix_dict_analysis.json')
            view_kline_data('./intraday_kline_data.json', symbol_to_group)
        elif choice == '0':
            main_menu()
        else:
            print("無效的選項，請重新輸入")

# ------------------ 更新K線數據：更新數據 ------------------
def update_kline_data():
    client, api_key = init_fugle_client()
    matrix_dict_analysis = load_matrix_dict_analysis()
    if not matrix_dict_analysis:
        print("沒有任何族群資料，請先管理族群。")
        return

    print("正在更新處置股清單...")
    fetch_disposition_stocks(client, matrix_dict_analysis)
    print("處置股清單已更新。")

    disposition_stocks = load_disposition_stocks()
    symbols_to_analyze = [sym for group in matrix_dict_analysis.values() for sym in group if sym not in disposition_stocks]

    # ===== ① 更新日 K 線資料 =====
    print("✅ 開始更新日K線數據至 daily_kline_data.json...")

    existing_daily_kline_data = {}
    if os.path.exists('daily_kline_data.json'):
        with open('daily_kline_data.json', 'r', encoding='utf-8') as f:
            try:
                existing_daily_kline_data = json.load(f)
            except json.JSONDecodeError:
                existing_daily_kline_data = {}
    else:
        print("⚠️ daily_kline_data.json 不存在，將建立新檔案。")

    data_is_same = True
    max_symbols_to_fetch = 20
    symbols_fetched = 0
    initial_api_count = 0

    for symbol in symbols_to_analyze[:max_symbols_to_fetch]:
        if initial_api_count >= 55:
            print("已達到55次API請求，休息1分鐘...")
            time_module.sleep(60)
            initial_api_count = 0

        daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
        initial_api_count += 1

        if daily_kline_df.empty:
            print(f"❌ 無法取得 {symbol} 的日K數據，跳過。")
            continue

        daily_kline_data = daily_kline_df.to_dict(orient='records')
        existing_data = existing_daily_kline_data.get(symbol)
        if existing_data != daily_kline_data:
            data_is_same = False
            print(f"{symbol} 的日K數據與現有資料不同，將更新資料。")
            existing_daily_kline_data[symbol] = daily_kline_data
        else:
            print(f"{symbol} 的日K數據與現有資料相同，跳過更新。")
        symbols_fetched += 1

    if not data_is_same:
        print("🔄 發現資料有異動，開始更新剩餘股票...")
        remaining_symbols = symbols_to_analyze[max_symbols_to_fetch:]
        for symbol in remaining_symbols:
            if initial_api_count >= 55:
                print("已達到55次API請求，休息1分鐘...")
                time_module.sleep(60)
                initial_api_count = 0

            daily_kline_df = fetch_daily_kline_data(client, symbol, days=2)
            initial_api_count += 1

            if daily_kline_df.empty:
                print(f"❌ 無法取得 {symbol} 的日K數據，跳過。")
                continue

            daily_kline_data = daily_kline_df.to_dict(orient='records')
            existing_data = existing_daily_kline_data.get(symbol)
            if existing_data != daily_kline_data:
                print(f"{symbol} 的日K數據與現有資料不同，將更新資料。")
                existing_daily_kline_data[symbol] = daily_kline_data
            else:
                print(f"{symbol} 的日K數據與現有資料相同，跳過更新。")

    with open('daily_kline_data.json', 'w', encoding='utf-8') as f:
        json.dump(existing_daily_kline_data, f, indent=4, ensure_ascii=False)

    print("✅ 日K線數據已寫入 daily_kline_data.json。")

    # ===== ② 更新一分 K 線資料 =====
    print("✅ 開始更新一分K線資料至 intraday_kline_data.json...")

    def get_recent_trading_day():
        today = datetime.now().date()
        now_time = datetime.now().time()
        market_open = datetime.strptime("09:00", "%H:%M").time()
        market_close = datetime.strptime("13:30", "%H:%M").time()

        def last_friday(date):
            while date.weekday() != 4:
                date -= timedelta(days=1)
            return date

        weekday = today.weekday()

        if weekday == 5:  # Saturday
            return last_friday(today)
        elif weekday == 6:  # Sunday
            return last_friday(today)
        elif weekday == 0:  # Monday
            if now_time < market_open:
                return last_friday(today)
            elif market_open <= now_time <= market_close:
                return last_friday(today)
            else:
                return today
        else:  # Tuesday to Friday
            if now_time < market_open:
                return today - timedelta(days=1)
            elif market_open <= now_time <= market_close:
                return today - timedelta(days=1)
            else:
                return today

    intraday_kline_data = {}
    count = 0

    trading_day = get_recent_trading_day().strftime('%Y-%m-%d')
    print(f"📅 本次一分K更新使用交易日: {trading_day}")

    for symbol in symbols_to_analyze:
        if count >= 55:
            print("已達到55次API請求，休息1分鐘...")
            time_module.sleep(60)
            count = 0

        daily_data = existing_daily_kline_data.get(symbol, [])
        if len(daily_data) < 2:
            print(f"{symbol} 日K資料不足，無法判斷昨收，跳過。")
            continue
        yesterday_close_price = daily_data[1].get('close', 0)
        intraday_df = fetch_intraday_data(
            client=client,
            symbol=symbol,
            trading_day=trading_day,
            yesterday_close_price=yesterday_close_price,
            start_time="09:00",
            end_time="13:30"
        )
        count += 1

        if intraday_df.empty:
            print(f"無法取得 {symbol} 的一分K數據，跳過。")
            continue
        
        updated_records = []
        records = intraday_df.to_dict(orient='records')
        for i, candle in enumerate(records): 
            updated_candle = calculate_2min_pct_increase_and_highest(candle, records[:i])
            updated_records.append(updated_candle)
        intraday_df = pd.DataFrame(updated_records)
        intraday_kline_data[symbol] = intraday_df.to_dict(orient='records')
        print(f"{symbol} 的一分K資料已加入。")

    intraday_kline_data_str = convert_datetime_to_str(intraday_kline_data)
    with open('intraday_kline_data.json', 'w', encoding='utf-8') as f:
        json.dump(intraday_kline_data_str, f, indent=4, ensure_ascii=False, default=str)
    print("✅ 一分K線資料已寫入 intraday_kline_data.json。")

    consolidate_and_save_stock_symbols()
    print("✅ 股票代號已統整並儲存至 nb_matrix_dict.json。")


# ------------------ 更新K線數據：查看數據 ------------------
def load_symbol_to_group(matrix_path):
    """
    從 matrix_dict_analysis.json 讀取族群資料，並轉換成 symbol -> group 字典
    """
    with open(matrix_path, 'r', encoding='utf-8') as f:
        matrix_dict = json.load(f)
    
    symbol_to_group = {}
    for group, symbols in matrix_dict.items():
        for symbol in symbols:
            symbol_to_group[symbol] = group
    
    return symbol_to_group

# ➡️ 加：設定中文顯示
plt.rcParams['font.family'] = 'Microsoft JhengHei'

def view_kline_data(json_path, symbol_to_group):
    """
    查看盤中K線數據，依族群分類並繪製標準化close走勢圖
    - 使用Z-score標準化
    - 自動處理中文顯示
    - 指定時間格式避免警告
    """
    # 讀取JSON
    if not os.path.exists(json_path):
        raise FileNotFoundError(f"找不到檔案：{json_path}")
        
    with open(json_path, 'r', encoding='utf-8') as f:
        raw_data = json.load(f)
    
    # 整理成 symbol -> DataFrame
    stock_data = {}
    for symbol, records in raw_data.items():
        df = pd.DataFrame(records)
        if 'time' in df.columns and 'close' in df.columns and 'date' in df.columns:
            # 先合併日期和時間
            df['datetime'] = pd.to_datetime(df['date'] + ' ' + df['time'], format="%Y-%m-%d %H:%M:%S")
            df = df.sort_values(by='datetime')
            stock_data[symbol] = df
        else:
            print(f"股票 {symbol} 缺少必要欄位，略過。")
    
    # 按族群分組
    group_to_stocks = {}
    for symbol, group in symbol_to_group.items():
        if symbol in stock_data:
            group_to_stocks.setdefault(group, []).append(symbol)
    
    all_figures = []
    
    # 為每個族群繪圖
    for group, symbols in group_to_stocks.items():
        fig, ax = plt.subplots(figsize=(12, 6))
        
        for symbol in symbols:
            df = stock_data[symbol]
            close = df['close']
            close_z = (close - close.mean()) / close.std() if close.std() != 0 else close - close.mean()
            ax.plot(df['datetime'], close_z, label=symbol)
        
        ax.set_title(f"{group} 族群標準化收盤價走勢")
        ax.set_xlabel("時間")
        ax.set_ylabel("標準化收盤價 (Z-score)")
        ax.legend()
        ax.grid(True)
        all_figures.append(fig)  # 存起來

    # 🔥 畫完所有族群圖之後，再一次顯示所有圖
    plt.show()




# ------------------ 查詢處置股 ------------------
def display_disposition_stocks():
    disposition_file = 'Disposition.json'
    try:
        with open(disposition_file, 'r', encoding='utf-8') as f:
            disposition_data = json.load(f)
            if isinstance(disposition_data, list):
                stock_codes = disposition_data
            elif isinstance(disposition_data, dict):
                stock_codes = disposition_data.get("stock_codes", [])
            else:
                print(f"錯誤：{disposition_file} 文件格式不正確。")
                return
    except FileNotFoundError:
        print(f"錯誤：無法找到 {disposition_file} 文件。")
        return
    except json.JSONDecodeError:
        print(f"錯誤：{disposition_file} 文件格式不正確。")
        return

    if not stock_codes:
        print(f"{disposition_file} 中沒有任何股票代號。")
        return

    items_per_page = 10
    total_items = len(stock_codes)
    total_pages = (total_items + items_per_page - 1) // items_per_page
    current_page = 1

    while True:
        start_idx = (current_page - 1) * items_per_page
        end_idx = start_idx + items_per_page
        page_items = stock_codes[start_idx:end_idx]

        print("\n" + "=" * 50)
        print(f"{disposition_file} 股票代號列表 - 第 {current_page} 頁 / 共 {total_pages} 頁")
        print("=" * 50)
        for idx, code in enumerate(page_items, start=1 + start_idx):
            print(f"{idx}. {code}")
        print("=" * 50)
        if total_pages == 1:
            print("已顯示所有股票代號。")
            break

        print("導航選項：")
        if current_page > 1:
            print("P - 上一頁")
        if current_page < total_pages:
            print("N - 下一頁")
        print("0 - 返回主選單")

        choice = input("請輸入選項（N/P/0）：").strip().upper()

        if choice == 'N' and current_page < total_pages:
            current_page += 1
        elif choice == 'P' and current_page > 1:
            current_page -= 1
        elif choice == '0':
            break
        else:
            print("無效的選項，請重新輸入。")




# ------------------ 工具程式 ------------------

# 計算三分鐘漲幅與當日最高價
def calculate_2min_pct_increase_and_highest(new_candle, existing_candles):
    new_candle['2min_pct_increase'] = 0.0
    new_candle['highest'] = new_candle.get('high', 0)

    if not existing_candles:
        return new_candle

    all_candles = existing_candles + [new_candle]
    num_existing_candles = len(existing_candles)

    # --- 計算 2分鐘漲幅 ---
    if num_existing_candles < 2:
        relevant_candles = all_candles
    else:
        relevant_candles = existing_candles[-1:] + [new_candle]  # 最近兩根

    rise_values = [float(c.get('rise', 0.0)) for c in relevant_candles if c.get('rise') is not None]

    if len(rise_values) >= 2:
        first = rise_values[0]
        last = rise_values[-1]
        r_max = max(rise_values)
        r_min = min(rise_values)

        if last >= first:
            pct_increase = r_max - r_min
        else:
            pct_increase = r_min - r_max

        new_candle['2min_pct_increase'] = round(pct_increase, 2)

    # --- 計算 highest ---
    prev_highest = max([c.get('highest', 0) for c in existing_candles])
    new_candle['highest'] = max(prev_highest, new_candle.get('high', 0))

    return new_candle

# 計算兩分鐘漲幅
def calculate_2min_pct_increase(new_candle, existing_candles):
    new_candle['2min_pct_increase'] = 0.0
    all_candles = existing_candles + [new_candle]
    num_existing_candles = len(existing_candles)

    if num_existing_candles == 0:
        return new_candle

    if num_existing_candles >=1:
        relevant_candles = all_candles
    else:
        relevant_candles = existing_candles[-1:] + [new_candle]  # 最近 2 根

    # 改為使用 rise 欄位
    rise_values = [float(c.get('rise', 0.0)) for c in relevant_candles if c.get('rise') is not None]

    if len(rise_values) < 2:
        new_candle['2min_pct_increase'] = 0.0
    else:
        first = rise_values[0]
        last = rise_values[-1]
        r_max = max(rise_values)
        r_min = min(rise_values)

        if last >= first:
            pct_increase = r_max - r_min
        else:
            pct_increase = r_min - r_max

        new_candle['2min_pct_increase'] = round(pct_increase, 2)

    return new_candle

# 計算一分K線數據（回測、歷史行情用）
def fetch_intraday_data(client, symbol, trading_day, yesterday_close_price, start_time=None, end_time=None):
    try:
        symbol = str(symbol).strip()
        if not symbol:
            print(f"❌ 無效的 symbol: {symbol}")
            return pd.DataFrame()

        if isinstance(trading_day, str):
            trading_day_date = datetime.strptime(trading_day, '%Y-%m-%d').date()
        elif isinstance(trading_day, datetime):
            trading_day_date = trading_day.date()
        elif isinstance(trading_day, date):
            trading_day_date = trading_day
        else:
            print(f"❌ 無效 trading_day 類型：{type(trading_day)}，值：{trading_day}")
            return pd.DataFrame()

        today = datetime.now().date()
        if trading_day_date < today:
            end_time_str = "13:30"
        else:
            now = datetime.now()
            market_end = now.replace(hour=13, minute=30, second=0, microsecond=0)
            end_time_str = "13:30" if now > market_end else (now - timedelta(minutes=1)).replace(second=0, microsecond=0).strftime('%H:%M')

        _from = datetime.strptime(f"{trading_day} {start_time or '09:00'}", "%Y-%m-%d %H:%M")
        to = datetime.strptime(f"{trading_day} {end_time or end_time_str}", "%Y-%m-%d %H:%M")

        candles_rsp = client.stock.intraday.candles(
            symbol=symbol, timeframe='1',
            _from=_from.isoformat(), to=to.isoformat()
        )

        if not candles_rsp or not candles_rsp.get('data'):
            print(f"⚠️ API 無回傳資料：{candles_rsp}")
            return pd.DataFrame()

        candles_df = pd.DataFrame(candles_rsp['data'])
        if 'volume' not in candles_df.columns:
            print(f"⚠️ volume 欄位不存在！實際欄位：{candles_df.columns.tolist()}")
            return pd.DataFrame()

        candles_df['volume'] = pd.to_numeric(candles_df['volume'], errors='coerce')
        candles_df['datetime'] = pd.to_datetime(candles_df['date'], errors='coerce').dt.tz_localize(None).dt.floor('min')
        candles_df.set_index('datetime', inplace=True)

        original_df = candles_df.reset_index()[['datetime', 'volume']].rename(columns={'volume': 'orig_volume'})

        full_idx = pd.date_range(start=_from, end=to, freq='1min')
        candles_df = candles_df.reindex(full_idx)

        candles_df.reset_index(inplace=True)
        candles_df.rename(columns={'index': 'datetime'}, inplace=True)
        candles_df['date'] = candles_df['datetime'].dt.strftime('%Y-%m-%d')
        candles_df['time'] = candles_df['datetime'].dt.strftime('%H:%M:%S')

        candles_df = pd.merge(candles_df, original_df, how='left', on='datetime')
        candles_df['was_filled'] = candles_df['orig_volume'].isna()

        # ✅ 若找不到前一根K棒，補上昨日收盤價
        for col in ['open', 'high', 'low', 'close']:
            values = candles_df[col].to_numpy()
            last_valid_close = yesterday_close_price

            for i in range(len(values)):
                vol = candles_df.at[i, 'volume']
                close_val = candles_df.at[i, 'close']

                if vol > 0 and not pd.isna(close_val):
                    last_valid_close = close_val

                if pd.isna(values[i]) or vol == 0:
                    values[i] = last_valid_close

            candles_df[col] = values

        candles_df['volume'] = candles_df['orig_volume'].fillna(0)
        candles_df['symbol'] = symbol
        candles_df['昨日收盤價'] = yesterday_close_price
        candles_df['漲停價'] = truncate_to_two_decimals(calculate_limit_up_price(yesterday_close_price))
        candles_df[['symbol', '昨日收盤價', '漲停價']] = candles_df[['symbol', '昨日收盤價', '漲停價']].ffill().bfill()
        candles_df['rise'] = (candles_df['close'] - candles_df['昨日收盤價']) / candles_df['昨日收盤價'] * 100

        # ✅ 修正 highest = None 問題：補完 high 再計算 cummax
        candles_df['highest'] = candles_df['high'].cummax()
        candles_df['highest'] = candles_df['highest'].fillna(yesterday_close_price)

        return candles_df[['symbol', 'date', 'time', 'open', 'high', 'low',
                           'close', 'volume', '昨日收盤價', '漲停價', 'rise', 'highest']]

    except Exception as e:
        print(f"❌ 發生例外錯誤：{e}")
        traceback.print_exc()
        return pd.DataFrame()
    
def fetch_realtime_intraday_data(client, symbol, trading_day, yesterday_close_price, start_time=None, end_time=None):
    try:
        symbol = str(symbol).strip()
        if not symbol:
            print(f"❌ 無效的 symbol: {symbol}")
            return pd.DataFrame()

        # 處理 trading_day 參數
        if isinstance(trading_day, str):
            trading_day_date = datetime.strptime(trading_day, '%Y-%m-%d').date()
        elif isinstance(trading_day, datetime):
            trading_day_date = trading_day.date()
        elif isinstance(trading_day, date):
            trading_day_date = trading_day
        else:
            print(f"❌ 無效 trading_day 類型：{type(trading_day)}，值：{trading_day}")
            return pd.DataFrame()

        # 計算結束時間
        today = datetime.now().date()
        if trading_day_date < today:
            end_time_str = "13:30"
        else:
            now = datetime.now()
            market_end = now.replace(hour=13, minute=30, second=0, microsecond=0)
            end_time_str = "13:30" if now > market_end else (now - timedelta(minutes=1)).replace(second=0, microsecond=0).strftime('%H:%M')

        _from = datetime.strptime(f"{trading_day} {start_time or '09:00'}", "%Y-%m-%d %H:%M")
        to = datetime.strptime(f"{trading_day} {end_time or end_time_str}", "%Y-%m-%d %H:%M")

        candles_rsp = client.stock.intraday.candles(
            symbol=symbol, timeframe='1',
            _from=_from.isoformat(), to=to.isoformat()
        )

        if not candles_rsp or not candles_rsp.get('data'):
            print(f"⚠️ API 無回傳資料：{candles_rsp}")
            return pd.DataFrame()

        candles_df = pd.DataFrame(candles_rsp['data'])
        if 'volume' not in candles_df.columns:
            print(f"⚠️ volume 欄位不存在！實際欄位：{candles_df.columns.tolist()}")
            return pd.DataFrame()

        candles_df['volume'] = pd.to_numeric(candles_df['volume'], errors='coerce')
        candles_df['datetime'] = pd.to_datetime(candles_df['date'], errors='coerce').dt.tz_localize(None).dt.floor('min')
        candles_df.set_index('datetime', inplace=True)

        original_df = candles_df.reset_index()[['datetime', 'volume']].rename(columns={'volume': 'orig_volume'})

        full_idx = pd.date_range(start=_from, end=to, freq='1min')
        candles_df = candles_df.reindex(full_idx)

        candles_df.reset_index(inplace=True)
        candles_df.rename(columns={'index': 'datetime'}, inplace=True)
        candles_df['date'] = candles_df['datetime'].dt.strftime('%Y-%m-%d')
        candles_df['time'] = candles_df['datetime'].dt.strftime('%H:%M:%S')

        candles_df = pd.merge(candles_df, original_df, how='left', on='datetime')
        candles_df['was_filled'] = candles_df['orig_volume'].isna()

        # ✅ 使用向量化補值取代 iterrows，效率提升
        filled = candles_df['was_filled'].to_numpy()
        close = candles_df['close'].to_numpy()

        last_valid = yesterday_close_price
        for i in range(len(close)):
            if candles_df.at[i, 'volume'] > 0:
                if not pd.isna(close[i]):
                    last_valid = close[i]
            if filled[i]:
                close[i] = last_valid

        for col in ['open', 'high', 'low', 'close']:
            values = candles_df[col].to_numpy()
            last_valid_close = None
            for i in range(len(values)):
                vol = candles_df.at[i, 'volume']
                if vol > 0 and not pd.isna(candles_df.at[i, 'close']):
                    last_valid_close = candles_df.at[i, 'close']
                if pd.isna(values[i]) or vol == 0:
                    if last_valid_close is not None:
                        values[i] = last_valid_close
            candles_df[col] = values

        # volume 本身補零
        candles_df['volume'] = candles_df['orig_volume'].fillna(0)

        # ✅ 專屬即時補值：volume==0 時讀取 auto_intraday.json 最後一筆 close
        try:
            with open("auto_intraday.json", "r", encoding="utf-8") as f:
                auto_data = json.load(f)
            last_close = None
            if symbol in auto_data and isinstance(auto_data[symbol], list) and auto_data[symbol]:
                last_record = auto_data[symbol][-1]
                last_close = last_record.get("close")

            for i in range(len(candles_df)):
                if candles_df.at[i, "volume"] == 0:
                    if last_close is not None:
                        for col in ["open", "high", "low", "close"]:
                            candles_df.at[i, col] = last_close
                        '''
                        print(f"📌 [補值] {symbol} {candles_df.at[i, 'time']} volume=0，使用 auto_intraday.json 最後一筆 close={last_close}")
                        '''
                    else:
                        print(f"⚠️ [補值失敗] {symbol} {candles_df.at[i, 'time']} volume=0，但無法從 auto_intraday.json 取得 close 值")
        except Exception as e:
            print(f"❌ 讀取 auto_intraday.json 時發生錯誤：{e}")

        candles_df['symbol'] = symbol
        candles_df['昨日收盤價'] = yesterday_close_price
        candles_df['漲停價'] = truncate_to_two_decimals(calculate_limit_up_price(yesterday_close_price))
        candles_df[['symbol', '昨日收盤價', '漲停價']] = candles_df[['symbol', '昨日收盤價', '漲停價']].ffill().bfill()
        candles_df['rise'] = (candles_df['close'] - candles_df['昨日收盤價']) / candles_df['昨日收盤價'] * 100
        candles_df['highest'] = candles_df['high'].cummax()

        return candles_df[[ 'symbol', 'date', 'time', 'open', 'high', 'low',
                            'close', 'volume', '昨日收盤價', '漲停價', 'rise', 'highest']]

    except Exception as e:
        print(f"❌ 發生例外錯誤：{e}")
        traceback.print_exc()
        return pd.DataFrame()


# 計算日K線數據
def fetch_daily_kline_data(client, symbol, days=2):
    end_date = get_recent_trading_day()
    start_date = end_date - timedelta(days=days)
    start_date_str = start_date.strftime('%Y-%m-%d')
    end_date_str = end_date.strftime('%Y-%m-%d')

    print(f"正在取得 {symbol} 從 {start_date_str} 到 {end_date_str} 的日K數據...")

    try:
        data = client.stock.historical.candles(symbol=symbol, from_=start_date_str, to=end_date_str)
        if 'data' in data and data['data']:
            daily_kline_df = pd.DataFrame(data['data'])
            return daily_kline_df
        else:
            print(f"無法取得 {symbol} 的日K數據：API 回應中不包含 'data' 欄位或 'data' 為空")
            return pd.DataFrame()
    except Exception as e:
        print(f"無法取得 {symbol} 的日K數據：{e}")
        return pd.DataFrame()

# 計算最近交易日
def get_recent_trading_day():
    today = datetime.now().date()
    now_time = datetime.now().time()
    market_open = datetime.strptime("09:00", "%H:%M").time()
    market_close = datetime.strptime("13:30", "%H:%M").time()

    def last_friday(date):
        """往回找到最近的星期五"""
        while date.weekday() != 4:
            date -= timedelta(days=1)
        return date

    weekday = today.weekday()  # 0=Monday, 6=Sunday

    if weekday == 5:  # Saturday
        return last_friday(today)
    elif weekday == 6:  # Sunday
        return last_friday(today)
    elif weekday == 0:  # Monday
        if now_time < market_open:
            return last_friday(today)  # 週一早上9點前，還是拿上週五
        elif market_open <= now_time <= market_close:
            return last_friday(today)  # 盤中也拿上週五
        else:
            return today  # 盤後拿今天
    else:  # Tuesday to Friday
        if now_time < market_open:
            return today - timedelta(days=1)
        elif market_open <= now_time <= market_close:
            return today - timedelta(days=1)
        else:
            return today

# 設定選單工具
def save_settings():
    with open('settings.json', 'w', encoding='utf-8') as f:
        json.dump({
            'capital_per_stock': capital_per_stock,
            'transaction_fee': transaction_fee,
            'transaction_discount': transaction_discount,
            'trading_tax': trading_tax,
            'below_50': below_50,
            'price_gap_50_to_100': price_gap_50_to_100,
            'price_gap_100_to_500': price_gap_100_to_500,
            'price_gap_500_to_1000': price_gap_500_to_1000,
            'price_gap_above_1000': price_gap_above_1000,
            'allow_reentry_after_stop_loss': allow_reentry_after_stop_loss
        }, f, indent=4)

def load_settings():
    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global below_50, price_gap_50_to_100, price_gap_100_to_500, price_gap_500_to_1000, price_gap_above_1000
    global allow_reentry_after_stop_loss
    if os.path.exists('settings.json'):
        with open('settings.json', 'r', encoding='utf-8') as f:
            settings = json.load(f)
            capital_per_stock = settings.get('capital_per_stock', 0)
            transaction_fee = settings.get('transaction_fee', 0)
            transaction_discount = settings.get('transaction_discount', 0)
            trading_tax = settings.get('trading_tax', 0)
            below_50 = settings.get('below_50', 0)
            price_gap_50_to_100 = settings.get('price_gap_50_to_100', 0)
            price_gap_100_to_500 = settings.get('price_gap_100_to_500', 0)
            price_gap_500_to_1000 = settings.get('price_gap_500_to_1000', 0)
            price_gap_above_1000 = settings.get('price_gap_above_1000', 0)
            allow_reentry_after_stop_loss = settings.get('allow_reentry_after_stop_loss', False)
    else:
        capital_per_stock = 1000
        transaction_fee = 0.1425
        transaction_discount = 20.0
        trading_tax = 0.15
        below_50 = 500
        price_gap_50_to_100 = 1000
        price_gap_100_to_500 = 2000
        price_gap_500_to_1000 = 3000
        price_gap_above_1000 = 5000
        allow_reentry_after_stop_loss = False

def consolidate_and_save_stock_symbols():
    # 只讀 matrix_dict_analysis.json，直接當作最終清單
    matrix_dict_analysis = load_matrix_dict_analysis()
    if not matrix_dict_analysis:
        print("matrix_dict_analysis.json 檔案不存在或為空，無法進行統整")
        return

    # 直接以原始群組定義建立 consolidated_symbols
    nb_matrix_dict = {
        "consolidated_symbols": matrix_dict_analysis
    }

    save_nb_matrix_dict(nb_matrix_dict)


# ------------------ 計算回測函數相似度 ------------------
def calculate_group_kline_data(
    leader: str,
    candidates: list[str],
    intraday_file: str,
    group_symbols: list[str],
    wait_end_time: datetime,
    similarity_threshold: float = 0.4,
    close_weight: float = 2.0
) -> list[str]:
    """
    計算「leader」與 group_symbols（除 leader 外）在 09:00～(wait_end_time - 1min) 之間
    的 K 線相似度，並把過程 dump 到 debug JSON。
    """

    # 1. 讀 intraday_kline_data.json
    with open(intraday_file, 'r', encoding='utf-8') as f:
        intraday_data = json.load(f)

    # 2. 定義時間範圍
    start_time = time(9, 0)
    end_dt     = wait_end_time - timedelta(minutes=1)
    end_time   = end_dt.time()
    print(f"[SIM DEBUG] 取用 K 線時間範圍：{start_time} ～ {end_time}")

    # 3. 取出 leader & group_symbols 的 DataFrame
    def df_of(sym):
        recs = intraday_data.get(sym, [])
        df = pd.DataFrame(recs)
        if df.empty:
            return df
        # 轉成 time 物件
        df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S').dt.time
        mask = (df['time'] >= start_time) & (df['time'] <= end_time)
        return df.loc[mask].sort_values('time').reset_index(drop=True)

    df_lead = df_of(leader)

    # 把 time 轉回字串，才不會 dump 時失敗
    df_lead_str = df_lead.copy()
    if not df_lead_str.empty:
        df_lead_str['time'] = df_lead_str['time'].astype(str)
    leader_klines = df_lead_str.to_dict('records')

    # 4. 計算所有 group_symbols（剔除 leader） vs leader 的相似度
    sim_results      = []
    candidate_klines = {}
    for sym in group_symbols:
        if sym == leader:
            continue
        df_cand = df_of(sym)
        # candidate 也同樣轉回字串
        df_cand_str = df_cand.copy()
        if not df_cand_str.empty:
            df_cand_str['time'] = df_cand_str['time'].astype(str)
        candidate_klines[sym] = df_cand_str.to_dict('records')

        if df_lead.empty or df_cand.empty:
            sim = None
        else:
            merged = pd.merge(df_lead, df_cand, on='time', suffixes=('_L','_C'))
            if merged.empty:
                sim = None
            else:
                # Z-score 標準化
                for col in ['open','high','low','close','volume']:
                    merged[f'{col}_L_z'] = (
                        merged[f'{col}_L'] - merged[f'{col}_L'].mean()
                    ) / merged[f'{col}_L'].std(ddof=0)
                    merged[f'{col}_C_z'] = (
                        merged[f'{col}_C'] - merged[f'{col}_C'].mean()
                    ) / merged[f'{col}_C'].std(ddof=0)

                dist = np.sqrt(
                    (merged['open_L_z']   - merged['open_C_z'])**2 +
                    (merged['high_L_z']   - merged['high_C_z'])**2 +
                    (merged['low_L_z']    - merged['low_C_z'])**2 +
                    close_weight*(merged['close_L_z'] - merged['close_C_z'])**2 +
                    (merged['volume_L_z'] - merged['volume_C_z'])**2
                ).mean()
                sim = 1 / (1 + dist)

        sim_results.append({'symbol': sym, 'similarity': sim})

    # 5. 產出符合門檻的股票清單
    sim_ok = [
        r['symbol']
        for r in sim_results
        if r['similarity'] is not None and r['similarity'] >= similarity_threshold
    ]

    # 6. Debug：把所有相似度結果與 K 線資料存到 JSON
    debug_info = {
        'leader':               leader,
        'group_symbols':        group_symbols,
        'time_range':           f"{start_time} ~ {end_time}",
        'similarity_threshold': similarity_threshold,
        'leader_klines':        leader_klines,
        'candidate_klines':     candidate_klines,
        'sim_results':          sim_results,
        'sim_ok':               sim_ok
    }
    with open('similarity_debug_full.json', 'w', encoding='utf-8') as f:
        json.dump(debug_info, f, ensure_ascii=False, indent=2)

    return sim_ok

# ------------------ 計算交易函數相似度 ------------------
def calculate_realtime_kline_data(
    leader: str,
    candidates: list[str],
    intraday_file: str,
    wait_end_time: datetime,
    similarity_threshold: float = 0.4,
    close_weight: float = 2.0
) -> list[str]:
    """
    跟 calculate_group_kline_data 幾乎一樣，
    只是資料來源改成 auto_intraday.json、輸出檔名改為 similarity_realtime_debug_full.json
    """
    import json
    from datetime import timedelta, time

    # 1. 讀 auto_intraday.json
    with open(intraday_file, 'r', encoding='utf-8') as f:
        intraday_data = json.load(f)

    # 2. 時間範圍：09:00 ～ wait_end_time - 1 分鐘
    start_time = time(9, 0)
    end_dt     = wait_end_time - timedelta(minutes=1)
    end_time   = end_dt.time()
    print(f"[RT-SIM DEBUG] 取用 K 線時間範圍：{start_time} ～ {end_time}")

    # 3. 取 DataFrame，切到時間區間
    def df_of(sym):
        recs = intraday_data.get(sym, [])
        df = pd.DataFrame(recs)
        if df.empty:
            return df
        df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S').dt.time
        return (
            df[(df['time'] >= start_time) & (df['time'] <= end_time)]
            .sort_values('time')
            .reset_index(drop=True)
        )

    df_lead = df_of(leader)
    leader_klines = df_lead.to_dict('records')

    sim_results = []
    candidate_klines = {}
    for sym in candidates:
        df_cand = df_of(sym)
        candidate_klines[sym] = df_cand.to_dict('records')

        merged = pd.merge(df_lead, df_cand, on='time', suffixes=('_L','_C'))
        if merged.empty:
            sim = None
        else:
            for col in ['open','high','low','close','volume']:
                merged[f'{col}_L_z'] = (
                    merged[f'{col}_L'] - merged[f'{col}_L'].mean()
                ) / merged[f'{col}_L'].std(ddof=0)
                merged[f'{col}_C_z'] = (
                    merged[f'{col}_C'] - merged[f'{col}_C'].mean()
                ) / merged[f'{col}_C'].std(ddof=0)
            dist = np.sqrt(
                (merged['open_L_z']   - merged['open_C_z'])**2 +
                (merged['high_L_z']   - merged['high_C_z'])**2 +
                (merged['low_L_z']    - merged['low_C_z'])**2 +
                close_weight*(merged['close_L_z'] - merged['close_C_z'])**2 +
                (merged['volume_L_z'] - merged['volume_C_z'])**2
            ).mean()
            sim = 1 / (1 + dist)
        sim_results.append({'symbol': sym, 'similarity': sim})

    sim_ok = [
        r['symbol']
        for r in sim_results
        if r['similarity'] is not None and r['similarity'] >= similarity_threshold
    ]

    # 6. 將所有 debug 資料寫出
    debug_info = {
        'leader':              leader,
        'time_range':          f"{start_time} ~ {end_time}",
        'similarity_threshold': similarity_threshold,
        'leader_klines':       leader_klines,
        'candidate_klines':    candidate_klines,
        'sim_results':         sim_results,
        'sim_ok':              sim_ok
    }
    with open('similarity_realtime_debug_full.json', 'w', encoding='utf-8') as f:
        json.dump(debug_info, f, ensure_ascii=False, indent=2)

    return sim_ok


def calculate_kline_similarity(stock_data_list, close_weight=2):
    """
    計算股票之間的K線相似度（含成交量，強化close重要性）
    
    參數：
    - stock_data_list: 每支股票的DataFrame列表（每個DataFrame包含至少open/high/low/close/volume欄位）
    - close_weight: close欄位在計算距離時的權重（預設為2倍）
    
    返回：
    - 相似度DataFrame（包含stock1、stock2、similarity_score）
    """
    similarity_results = []
    num_stocks = len(stock_data_list)
    
    for i in range(num_stocks):
        stock1 = stock_data_list[i]
        if 'symbol' not in stock1.columns:
            raise KeyError("DataFrame does not contain 'symbol' column.")
        symbol1 = stock1['symbol'].iloc[0]
        
        for j in range(i + 1, num_stocks):
            stock2 = stock_data_list[j]
            if 'symbol' not in stock2.columns:
                raise KeyError("DataFrame does not contain 'symbol' column.")
            symbol2 = stock2['symbol'].iloc[0]
            
            if symbol1 != symbol2:
                merged_df = pd.merge(stock1, stock2, on='time', suffixes=('_1', '_2'))
                merged_df['昨日收盤價_2'] = merged_df['昨日收盤價_2'].ffill().bfill()
                
                # 確認需要的欄位存在
                required_cols = ['open', 'high', 'low', 'close', 'volume']
                if not all(f'{col}_1' in merged_df.columns and f'{col}_2' in merged_df.columns for col in required_cols):
                    print(f"股票 {symbol1} 或 {symbol2} 缺少必要欄位，跳過相似度計算。")
                    continue

                # Z-score 標準化
                for col in required_cols:
                    merged_df[f'{col}_1_z'] = (merged_df[f'{col}_1'] - merged_df[f'{col}_1'].mean()) / merged_df[f'{col}_1'].std()
                    merged_df[f'{col}_2_z'] = (merged_df[f'{col}_2'] - merged_df[f'{col}_2'].mean()) / merged_df[f'{col}_2'].std()

                # 計算歐式距離，close欄位有加權
                distance = np.sqrt(
                    (merged_df['open_1_z'] - merged_df['open_2_z']) ** 2 +
                    (merged_df['high_1_z'] - merged_df['high_2_z']) ** 2 +
                    (merged_df['low_1_z'] - merged_df['low_2_z']) ** 2 +
                    close_weight * (merged_df['close_1_z'] - merged_df['close_2_z']) ** 2 +
                    (merged_df['volume_1_z'] - merged_df['volume_2_z']) ** 2
                ).mean()

                similarity_score = 1 / (1 + distance)
                
                if similarity_score >= 0.4:
                    result = {
                        'stock1': symbol1,
                        'stock2': symbol2,
                        'similarity_score': similarity_score
                    }
                    similarity_results.append(result)

    if not similarity_results:
        print("沒有找到相似度大於等於 0.4 的結果")
        return pd.DataFrame(columns=['stock1', 'stock2', 'similarity_score'])
    
    similarity_df = pd.DataFrame(similarity_results)
    similarity_df = similarity_df.sort_values(by='similarity_score', ascending=False).reset_index(drop=True)
    return similarity_df

def calculate_limit_up_price(close_price):
    limit_up = close_price * 1.10
    if limit_up < 10:
        price_unit = 0.01
    elif limit_up < 50:
        price_unit = 0.05
    elif limit_up < 100:
        price_unit = 0.1
    elif limit_up < 500:
        price_unit = 0.5
    elif limit_up < 1000:
        price_unit = 1
    else:
        price_unit = 5
    limit_up_price = (limit_up // price_unit) * price_unit
    return limit_up_price

def load_nb_matrix_dict():
    if os.path.exists('nb_matrix_dict.json'):
        with open('nb_matrix_dict.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        return {}
    
def ensure_continuous_time_series(df):
    df['date'] = pd.to_datetime(df['date'])
    df['time'] = pd.to_datetime(df['time'], format='%H:%M:%S').dt.time

    full_time_index = pd.date_range(start='09:00', end='13:30', freq='1min').time
    full_index = pd.MultiIndex.from_product([df['date'].unique(), full_time_index], names=['date', 'time'])

    df.set_index(['date', 'time'], inplace=True)
    df = df.reindex(full_index)
    df[['symbol', '昨日收盤價', '漲停價']] = df[['symbol', '昨日收盤價', '漲停價']].ffill().bfill()

    if 'high' not in df.columns:
        df['high'] = df['close']
    if 'low' not in df.columns:
        df['low'] = df['close']

    df['close'] = df['close'].ffill()
    df['close'] = df['close'].fillna(df['昨日收盤價'])
    df['open'] = df['open'].ffill()
    df['open'] = df['open'].fillna(df['close'])
    df['high'] = df['high'].ffill()
    df['high'] = df['high'].fillna(df['close'])
    df['low'] = df['low'].ffill()
    df['low'] = df['low'].fillna(df['close'])
    df['volume'] = df['volume'].fillna(0)

    if '2min_pct_increase' not in df.columns:
        df['2min_pct_increase'] = 0.0
    else:
        df['2min_pct_increase'] = df['2min_pct_increase'].fillna(0.0)

    df.reset_index(inplace=True)
    return df
        
def load_disposition_stocks():
    disposition_file = 'Disposition.json'
    try:
        with open(disposition_file, 'r', encoding='utf-8') as f:
            disposition_data = json.load(f)
            return disposition_data
    except FileNotFoundError:
        print(f"錯誤：無法找到 {disposition_file} 文件。")
        return []
    except json.JSONDecodeError:
        print(f"錯誤：{disposition_file} 文件格式不正確。")
        return []
    
def fetch_disposition_stocks(client, matrix_dict_analysis):
    disposition_stocks = []
    for group, stock_list in matrix_dict_analysis.items():
        for symbol in stock_list:
            try:
                ticker_data = client.stock.intraday.ticker(symbol=symbol)
                if ticker_data.get('isDisposition', False):
                    disposition_stocks.append(symbol)
            except Exception as e:
                print(f"獲取 {symbol} 的處置股狀態時發生錯誤: {e}")
    with open('Disposition.json', 'w', encoding='utf-8') as f:
        json.dump(disposition_stocks, f, indent=4, ensure_ascii=False)

def load_kline_data():
    daily_kline_data = {}
    intraday_kline_data = {}

    if os.path.exists('daily_kline_data.json'):
        with open('daily_kline_data.json', 'r', encoding='utf-8') as f:
            try:
                daily_kline_data = json.load(f)
                if not daily_kline_data:
                    print("日K線數據檔案為空，請先更新數據。")
            except json.JSONDecodeError:
                print("日K線數據檔案格式錯誤，請先更新數據。")

    if os.path.exists('intraday_kline_data.json'):
        with open('intraday_kline_data.json', 'r', encoding='utf-8') as f:
            try:
                intraday_kline_data = json.load(f)
                if not intraday_kline_data:
                    print("一分K線數據檔案為空，請先更新數據。")
            except json.JSONDecodeError:
                print("一分K線數據檔案格式錯誤，請先更新數據。")

    return daily_kline_data, intraday_kline_data

def load_symbols_to_analyze():
    nb_matrix_dict = load_nb_matrix_dict()
    consolidated_symbols = nb_matrix_dict.get("consolidated_symbols", {})
    symbols = []
    for group_symbols in consolidated_symbols.values():
        symbols.extend(group_symbols)
    disposition_stocks = load_disposition_stocks()
    symbols = [symbol for symbol in symbols if symbol not in disposition_stocks]
    return symbols

def load_group_symbols():
    if not os.path.exists('nb_matrix_dict.json'):
        print("nb_matrix_dict.json 文件不存在。")
        return {}
    with open('nb_matrix_dict.json', 'r', encoding='utf-8') as f:
        group_symbols = json.load(f)
    return group_symbols

# 把處置股從 nb_matrix_dict.json 剔除
def purge_disposition_from_nb(disposition_list, nb_path='nb_matrix_dict.json'):
    """
    disposition_list : List[str]  # 處置股代號清單
    nb_path          : str        # nb_matrix_dict 檔案路徑
    --------------
    讀取 nb_matrix_dict.json → consolidated_symbols
    若該股票代號出現在 disposition_list，便將其從對應族群移除。
    有異動才覆寫檔案。
    """
    if not os.path.exists(nb_path):
        print(f"找不到 {nb_path}，跳過處置股過濾。")
        return

    try:
        with open(nb_path, 'r', encoding='utf-8') as f:
            nb_dict = json.load(f)
    except json.JSONDecodeError:
        print(f"{nb_path} 格式錯誤，無法解析，跳過過濾。")
        return

    if 'consolidated_symbols' not in nb_dict or not isinstance(nb_dict['consolidated_symbols'], dict):
        print(f"{nb_path} 缺少 consolidated_symbols，跳過過濾。")
        return

    changed = False
    for grp, syms in nb_dict['consolidated_symbols'].items():
        # 原本可能有重複，先去重再過濾
        filtered = [s for s in dict.fromkeys(syms) if s not in disposition_list]
        if len(filtered) != len(syms):
            nb_dict['consolidated_symbols'][grp] = filtered
            changed = True

    # 若有異動，寫回檔案
    if changed:
        with open(nb_path, 'w', encoding='utf-8') as f:
            json.dump(nb_dict, f, ensure_ascii=False, indent=4)
        print(f"已從 {nb_path} 移除處置股：{', '.join(disposition_list)}")
    else:
        print("nb_matrix_dict.json 無需調整，未包含任何處置股。")

# 檢查盤中退出
def check_quit_flag_loop():
    while True:
        time_module.sleep(5)  # 每 5 秒檢查一次
        if quit_flag["quit"]:
            threading.Thread(target=show_exit_menu, daemon=True).start()
            quit_flag["quit"] = False

# 全域變數，用來記錄上一次存在於停損委託單中的股票代號
previous_stop_loss_codes = set()
open_positions: dict[str, dict] = {} # ‑ 只要有進場就寫入；平倉就刪除（盤中持倉表）。
triggered_limit_up_stocks: set[str] = set()

def monitor_stop_loss_orders():
    """
    每次呼叫時檢查 to.conditions 的內容，如果發現原本存在的停損委託單股票代號已不見，
    則檢查 allow_reentry_after_stop_loss 是否為 True，
    若是，則將該股票所屬族群的 in_position 設為 False（允許重入）。
    """
    global to, group_positions, previous_stop_loss_codes, allow_reentry_after_stop_loss

    # 取得目前停損委託單的股票代號集合
    if isinstance(to.conditions, dict):
        current_codes = set(to.conditions.keys())
    else:
        # 如果 to.conditions 不是字典，就嘗試從每個停損單物件中提取股票代號（依實際格式調整）
        current_codes = set()
        for cond in to.conditions:
            try:
                current_codes.add(cond.order_contract.code)
            except Exception as e:
                print(f"提取停損單代號時發生錯誤：{e}")

    # 與上一輪記錄比較，找出已移除的股票代號
    removed_codes = previous_stop_loss_codes - current_codes

    if removed_codes:
        if allow_reentry_after_stop_loss:
            nb_matrix_dict = load_nb_matrix_dict()  # 假設此函數能正確載入 nb_matrix_dict.json
            if "consolidated_symbols" in nb_matrix_dict:
                consolidated_symbols = nb_matrix_dict["consolidated_symbols"]
                for code in removed_codes:
                    # 尋找該股票所在的族群
                    for group, symbols in consolidated_symbols.items():
                        # 假設股票代號格式一致
                        if code in symbols:
                            if group in group_positions and group_positions[group] == "已進場":
                                group_positions[group] = False
                                print(f"停損觸發：股票 {code} 的停損委託單消失，將族群 {group} 的 in_position 設為 False")
            else:
                print("nb_matrix_dict 中缺少 'consolidated_symbols' 鍵，無法更新族群狀態")
        else:
            print("停損委託單消失，但停損再進場已關閉")
    else:
        print("監控中：目前未發現異常...")
        print("=" * 50)

    previous_stop_loss_codes = current_codes.copy()

def monitor_quit_key():
    """背景執行的 Q 鍵偵測器，按下 Q 將 quit_flag['quit'] 設為 True"""
    while True:
        if msvcrt.kbhit():
            key = msvcrt.getch().decode("utf-8").upper()
            if key == 'Q':
                quit_flag['quit'] = True

def show_exit_menu():
    """非阻塞地顯示退出平倉選單（實際平倉邏輯實作）"""
    def _menu():
        print("\n================ 手動退出選單 ================")
        print("1. 直接退出，不平倉")
        print("2. 平倉")
        print("0. 返回程式")
        choice = input("請輸入選項：").strip()
        if choice == "1":
            confirm = input("⚠️  確定不平倉直接退出？(Y/N)：").strip().upper()
            if confirm == "Y":
                os._exit(0)
                main_menu()

        elif choice == "2":
            while True:
                list_open_positions()
                print("\n平倉選項：1. 全部平倉  2. 選擇股票  0. 返回程式")
                sub = input("請輸入：").strip()
                if sub == "1":
                    exit_trade_live()
                    os._exit(0)
                    main_menu()
                elif sub == "2":
                    if not open_positions:
                        continue
                    code = input("輸入要平倉的股票代號：").strip()
                    if code in open_positions:
                        close_one_stock(code)
                    else:
                        print("代號不存在於持倉")
                    cont = input("已處理，繼續執行程式？(Y=繼續/N=退出)：").strip().upper()
                    if cont == "N":
                        os._exit(0)
                        main_menu()
                elif sub == "0":
                    break
        else:
            print("❌ 無效選項，繼續執行程式。")

    threading.Thread(target=_menu, daemon=True).start()
def save_matrix_dict(matrix_dict):
    with open('matrix_dict_analysis.json', 'w', encoding='utf-8') as f:
        json.dump(matrix_dict, f, indent=4, ensure_ascii=False)

def load_matrix_dict_analysis():
    if os.path.exists('matrix_dict_analysis.json'):
        with open('matrix_dict_analysis.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        print("matrix_dict_analysis.json 文件不存在。")
        return {}

def save_nb_matrix_dict(nb_matrix_dict):
    with open('nb_matrix_dict.json', 'w', encoding='utf-8') as f:
        json.dump(nb_matrix_dict, f, indent=4, ensure_ascii=False, default=str)

def initialize_stock_data(symbols_to_analyze, daily_kline_data, intraday_kline_data):
    stock_data_collection = {}
    for symbol in symbols_to_analyze:
        if symbol not in daily_kline_data or symbol not in intraday_kline_data:
            print(f"股票代號 {symbol} 的日 K 線或一分 K 線資料缺失，跳過。")
            continue
        daily_kline_df = pd.DataFrame(daily_kline_data[symbol])
        intraday_data = pd.DataFrame(intraday_kline_data[symbol])
        if intraday_data.empty:
            print(f"股票代號 {symbol} 的日內數據為空，跳過。")
            continue
        complete_df = ensure_continuous_time_series(intraday_data)
        complete_df = complete_df.drop(columns=['average'], errors='ignore')
        stock_data_collection[symbol] = complete_df
    return stock_data_collection

    
def truncate_to_two_decimals(value):
    if isinstance(value, float):
        return math.floor(value * 100) / 100
    return value

def save_auto_intraday_data(auto_intraday_data):
    try:
        with open('auto_intraday.json', 'wb') as f:
            f.write(orjson.dumps(auto_intraday_data, option=orjson.OPT_NON_STR_KEYS))
        print(f"{YELLOW}✅ 已儲存 auto_intraday.json{RESET}")
    except Exception as e:
        print(f"{YELLOW}❌ 儲存 auto_intraday.json 時發生錯誤：{e}{RESET}")

def initialize_triggered_limit_up(auto_intraday_data: dict):
    """
    掃描 auto_intraday.json 裡的歷史 1 分 K：
    只要曾出現『當根 high == 漲停價 且 前一根 high < 漲停價』
    就把該股票加入 triggered_limit_up_stocks，當天之後不再觸發。
    """
    now_txt = datetime.now().strftime("%H:%M")
    for sym, kbars in auto_intraday_data.items():
        for i in range(1, len(kbars)):
            prev, curr = kbars[i-1], kbars[i]
            if curr["high"] == curr["漲停價"] and prev["high"] < curr["漲停價"]:
                triggered_limit_up_stocks.add(sym)
                print(f"[{now_txt}] 股票 {sym} 曾經漲停觸發 → 加入已觸發清單")
                break  # 只要記一次就好

def convert_datetime_to_str(obj):
    if isinstance(obj, dict):
        return {k: convert_datetime_to_str(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [convert_datetime_to_str(element) for element in obj]
    elif isinstance(obj, (datetime, pd.Timestamp, time)):
        return obj.isoformat()
    else:
        return obj

# ------------------ 回測程式主程式 ------------------
def process_group_data(stock_data_collection, wait_minutes, hold_minutes,
                       matrix_dict_analysis, verbose=True):
    """
    === 回測函數 (Back-test)  ===
    - 同步 process_live_trading_logic 的四大邏輯：
      1. 拉高觸發：5-min 漲幅 ≥ 2% 且成交量 > 1.5×(09:00-09:02 平均量)
      2. 追蹤清單加入門檻：5-min 漲幅 ≥ 1.5%
      3. 漲停觸發：high == 漲停價 且 (前一根 high < 漲停價，09:00 例外)
      4. 等待期滿後的 eligible 篩選與進場、停損邏輯
    """

    # ────────── 0-A. 本地旗標初始化 ────────── #
    in_position         = False
    has_exited          = False
    current_position    = None
    stop_loss_triggered = False
    final_check_active  = False        # 回測版仍保留但目前未用
    final_check_count   = 0            # 〃
    hold_time           = 0

    # ────────── 0-B. 需要的全域設定 ────────── #
    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global price_gap_below_50, price_gap_50_to_100, price_gap_100_to_500
    global price_gap_500_to_1000, price_gap_above_1000
    global allow_reentry_after_stop_loss
    # --------------------------------------------------------------

    # ---------- 0-C. 開盤前三分鐘平均量 ---------- #
    FIRST3_AVG_VOL: dict[str, float] = {}
    for sym, df in stock_data_collection.items():
        first3 = df[df['time'].astype(str).isin(['09:00:00', '09:01:00', '09:02:00'])]
        FIRST3_AVG_VOL[sym] = first3['volume'].mean() if not first3.empty else 0

    # ---------- 0-D. 其他狀態變數 ---------- #
    message_log: list[tuple[str, str]] = []
    tracking_stocks: set[str] = set()
    leader                      = None
    leader_peak_rise            = None
    leader_rise_before_decline  = None
    in_waiting_period           = False
    waiting_time                = 0
    pull_up_entry               = False
    limit_up_entry              = False
    first_condition_one_time    = None

    # ---------- 0-E. 組 merge DataFrame ---------- #
    merged_df = None
    req_cols = ['time', 'rise', 'high', '漲停價',
                'close', '2min_pct_increase', 'volume']
    for sym, df in stock_data_collection.items():
        if not all(c in df.columns for c in req_cols):
            continue
        tmp = df[req_cols].copy()
        tmp = tmp.rename(columns={
            'rise':               f'rise_{sym}',
            'high':               f'high_{sym}',
            '漲停價':             f'limit_up_price_{sym}',
            'close':              f'close_{sym}',
            '2min_pct_increase':  f'2min_pct_increase_{sym}',
            'volume':             f'volume_{sym}'
        })
        merged_df = tmp if merged_df is None else pd.merge(
            merged_df, tmp, on='time', how='outer')

    if merged_df is None or merged_df.empty:
        print("無有效資料可回測")
        return None, None
    merged_df.sort_values('time', inplace=True, ignore_index=True)

    # ═══════════ 1. 逐分鐘主迴圈 ═══════════ #
    total_profit = total_profit_rate = total_trades = 0

    for _, row in merged_df.iterrows():
        current_time     = row['time']
        current_time_str = current_time.strftime('%H:%M:%S')

        # ── 1-1. 持倉期間：強制 / 時間平倉 / 條件停損 ── #
        if in_position and not has_exited:
            hold_time += 1

            # a) 13:30 強制
            if current_time_str == '13:30:00':
                profit, rate = exit_trade(
                    stock_data_collection[current_position['symbol']],
                    current_position['shares'],
                    current_position['entry_price'],
                    current_position['sell_cost'],
                    current_position['entry_fee'],
                    current_position['tax'],
                    message_log,
                    current_time, hold_time,
                    current_position['entry_time'],
                    use_f_exit=True
                )
                if profit is not None:
                    total_trades += 1
                    total_profit += profit
                    total_profit_rate += rate
                in_position = False
                has_exited  = True
                current_position = None
                continue

            # b) 持有分鐘到期
            if hold_minutes is not None and hold_time >= hold_minutes:
                profit, rate = exit_trade(
                    stock_data_collection[current_position['symbol']],
                    current_position['shares'],
                    current_position['entry_price'],
                    current_position['sell_cost'],
                    current_position['entry_fee'],
                    current_position['tax'],
                    message_log,
                    current_time, hold_time,
                    current_position['entry_time']
                )
                if profit is not None:
                    total_trades += 1
                    total_profit += profit
                    total_profit_rate += rate
                in_position = False
                has_exited  = True
                continue

            # c) 停損條件三
            sel_df  = stock_data_collection[current_position['symbol']]
            now_row = sel_df[sel_df['time'] == current_time]
            if not now_row.empty:
                h_now = truncate_to_two_decimals(now_row.iloc[0]['high'])
                thresh = truncate_to_two_decimals(
                    current_position['stop_loss_threshold'])
                if h_now >= thresh:
                    exit_price = thresh
                    exit_cost  = current_position['shares'] * exit_price * 1000
                    exit_fee   = int(exit_cost * (transaction_fee*0.01) *
                                     (transaction_discount*0.01))
                    profit = (current_position['sell_cost'] - exit_cost
                              - current_position['entry_fee'] - exit_fee
                              - current_position['tax'])
                    rate = (profit * 100) / (current_position['sell_cost']
                                              - current_position['entry_fee']
                                              - exit_fee)
                    message_log.append(
                        (current_time_str,
                         f"{Fore.RED}停損觸發，利潤 {int(profit)} 元 "
                         f"({rate:.2f}%){Style.RESET_ALL}")
                    )
                    total_trades += 1
                    total_profit += profit
                    total_profit_rate += rate
                    in_position = False
                    has_exited  = True
                    current_position = None
                    stop_loss_triggered = True
                    if not allow_reentry_after_stop_loss:
                        break
            continue  # 持倉時不再檢查新觸發

        # ── 1-2. 檢查觸發 (拉高/漲停) ── #
        trigger_list = []
        for sym in stock_data_collection.keys():
            pct  = row.get(f'2min_pct_increase_{sym}')
            vol  = row.get(f'volume_{sym}')
            high = row.get(f'high_{sym}')
            lup  = row.get(f'limit_up_price_{sym}')
            avgv = FIRST3_AVG_VOL.get(sym, 0)

            # 漲停
            hit_limit = False
            if high is not None and lup is not None and high == lup:
                if current_time_str == '09:00:00':
                    hit_limit = True
                else:
                    prev_time = (datetime.combine(date.today(), current_time)
                                 - timedelta(minutes=1)).time()
                    prev_high = stock_data_collection[sym].loc[
                        stock_data_collection[sym]['time'] == prev_time,
                        'high']
                    if prev_high.empty or prev_high.iloc[0] < lup:
                        hit_limit = True
            if hit_limit:
                trigger_list.append({'symbol': sym, 'condition': 'limit_up'})
                continue

            # 拉高
            if (pct is not None and pct >= 2
               and vol is not None and avgv and vol > 1.3*avgv):
                trigger_list.append({'symbol': sym, 'condition': 'pull_up'})

        # ── 1-3. 處理觸發結果 → 更新 tracking / leader / waiting ── #
        for item in trigger_list:
            sym  = item['symbol']
            cond = item['condition']
            if cond == 'limit_up':
                tracking_stocks.clear()
                tracking_stocks.add(sym)
                leader = sym
                in_waiting_period = True
                waiting_time = 0
                pull_up_entry  = False
                limit_up_entry = True
                first_condition_one_time = datetime.combine(date.today(), current_time)
                if verbose:
                    message_log.append(
                        (current_time_str,
                         f"{Fore.CYAN}{sym} 漲停觸發，開始等待{Style.RESET_ALL}"))
            else:  # pull_up
                if not pull_up_entry:
                    pull_up_entry = True
                    limit_up_entry = False
                    tracking_stocks.clear()
                    first_condition_one_time = datetime.combine(date.today(), current_time)
                tracking_stocks.add(sym)
                if verbose:
                    message_log.append(
                        (current_time_str,
                         f"{sym} 拉高觸發，加入追蹤"))

        # 追蹤清單擴充門檻 1.5%
        if pull_up_entry:
            for sym in stock_data_collection.keys():
                if sym in tracking_stocks:
                    continue
                pct = row.get(f'2min_pct_increase_{sym}')
                if pct is not None and pct >= 1.5:
                    tracking_stocks.add(sym)

        # ── 1-4. 領漲選擇與反轉偵測 ── #
        if pull_up_entry and tracking_stocks:
            # 選擇 rise 最大者
            max_sym, max_rise = None, None
            for sym in tracking_stocks:
                r = row.get(f'rise_{sym}')
                if r is not None and (max_rise is None or r > max_rise):
                    max_rise, max_sym = r, sym
            if leader != max_sym:
                if leader and verbose:
                    message_log.append(
                        (current_time_str,
                        f"{Fore.CYAN}領漲替換：{leader} → {max_sym}{Style.RESET_ALL}"))

                leader = max_sym
                leader_peak_rise = max_rise
                leader_rise_before_decline = max_rise

                # ✅ 領漲替換只更新，不進入等待
                in_waiting_period = False
                waiting_time = 0
                if verbose:
                    message_log.append(
                        (current_time_str,
                        f"{Fore.MAGENTA}🚀 領漲替換觸發，重新監控領漲{Style.RESET_ALL}"))
            # 反轉 → 進入等待
            if leader:
                h_now = row.get(f'high_{leader}')
                prev_time = (datetime.combine(date.today(), current_time)
                             - timedelta(minutes=1)).time()
                prev_row = stock_data_collection[leader][
                    stock_data_collection[leader]['time'] == prev_time]
                if not prev_row.empty:
                    h_prev = prev_row.iloc[0]['high']
                    if h_now <= h_prev and not in_waiting_period:
                        in_waiting_period = True
                        waiting_time = 0
                        leader_rise_before_decline = max_rise
                        if verbose:
                            message_log.append(
                                (current_time_str,
                                 f"領漲 {leader} 反轉，開始等待"))

        # ── 1-5. 等待時間計數 & 完成後篩選 eligible ── #
        if in_waiting_period:
            if waiting_time >= wait_minutes:
                # 等待完成，重置 flag
                in_waiting_period = False
                waiting_time      = 0

                # ─── 先找出這輪等待的 leader 屬於哪個群組 ───
                current_group = next(
                    (g for g, syms in matrix_dict_analysis.items() if leader in syms),
                    None
                )
                if current_group is None:
                    # 無法識別群組，跳過相似度過濾
                    filtered_stocks = set(tracking_stocks)
                else:
                    # 以相似度先篩選跟 leader 相近的股票
                    sim_ok = calculate_group_kline_data(
                        leader=leader,
                        candidates=list(tracking_stocks),
                        intraday_file="intraday_kline_data.json",
                        group_symbols=matrix_dict_analysis[current_group],
                        wait_end_time=datetime.combine(date.today(), current_time),
                        similarity_threshold=0.4,
                        close_weight=2.0
                    )
                    if verbose:
                        message_log.append(
                            (current_time_str,
                            f"[SIM] 相似度合格：{sim_ok}")
                        )
                    
                    # **新增 debug：確認 tracking_stocks 與 sim_ok**
                    message_log.append((current_time_str, f"[DEBUG] tracking_stocks = {tracking_stocks}"))
                    

                    filtered_stocks = set(sim_ok)

                    
                    # **立刻印出 filtered_stocks，確認真正帶進量能過濾的是哪些股票**  
                    if verbose:
                        message_log.append((current_time_str, f"[DEBUG] filtered_stocks = {filtered_stocks}"))
                    


                # ── 接著才進入原本的 eligible 篩選流程 ──
                def _vol_break(sym, join_time):
                    df   = stock_data_collection[sym]
                    avgv = FIRST3_AVG_VOL.get(sym, 0)
                    later = df[df['time'] >= join_time.time()]
                    return (later['volume'] >= 1.5*avgv).any()

                def _rise_peak_flat(sym: str, join_time: datetime) -> bool:
                    """
                    檢查從 join_time 開始算，sym 這支股票的 rise 是否「先見高點後不再創高」。
                    同時印出 debug 訊息：rise 序列、最高點、後續最大值，以及是否通過。
                    """
                    df = stock_data_collection[sym]
                    # 取出從 join_time 之後的所有 K 棒
                    sub = df[df['time'] >= join_time.time()]
                    if sub.empty:
                        if verbose:
                            message_log.append((current_time_str,
                                f"[DEBUG-PEAK] {sym} 無任何 join_time 之後的資料 → 不通過"))
                        return False

                    # 找出最高點
                    pkidx = sub['rise'].idxmax()
                    pkval = sub.loc[pkidx, 'rise']

                    # 後續所有點
                    later = sub.loc[pkidx+1:]
                    later_max = later['rise'].max() if not later.empty else None

                    # 判斷是否再創高
                    passed = (later_max is None) or (later_max <= pkval)

                    # 偵錯輸出
                    if verbose:
                        message_log.append((current_time_str,
                            f"[DEBUG-PEAK] {sym} rises={sub['rise'].tolist()}"))
                        message_log.append((current_time_str,
                            f"[DEBUG-PEAK] {sym} peak={pkval:.2f} at idx={pkidx}, "
                            f"later_max={later_max if later_max is not None else 'N/A'}, pass={passed}"))

                    return passed

                # 1. 列出量能突破門檻的股票
                passed_vol = [s for s in filtered_stocks if _vol_break(s, first_condition_one_time)]
                
                message_log.append((current_time_str, f"[DEBUG] 成交量通過: {passed_vol}"))
                
                
                # 2. 列出反轉通過的股票及峰頂時間
                passed_reversal = []
                for s in passed_vol:
                    df    = stock_data_collection[s]
                    sub   = df[df['time'] >= first_condition_one_time.time()]
                    pkidx = sub['rise'].idxmax()
                    pk_time = sub.loc[pkidx, 'time'].strftime('%H:%M:%S')
                    # 用 _rise_peak_flat 同時計算 passed
                    if _rise_peak_flat(s, first_condition_one_time):
                        passed_reversal.append((s, pk_time))
                message_log.append((current_time_str, f"[DEBUG] 反轉通過: {passed_reversal}"))

                # 3. 建立 eligible 並把反轉時間帶入 debug
                eligible = []
                for s, pk_time in passed_reversal:
                    rise_now = row[f"rise_{s}"]
                    price_now= row[f"close_{s}"]
                    if rise_now is None or not (-2 <= rise_now <= 6): continue
                    if price_now is None or price_now > capital_per_stock*1.5: continue
                    row_sym = stock_data_collection[s][stock_data_collection[s]['time']==current_time].iloc[0]
                    eligible.append({'symbol': s, 'rise': rise_now, 'row': row_sym, 'reversal_time': pk_time})
                    message_log.append((current_time_str,
                        f"[ELIGIBLE] {s} 峰頂 {pk_time} 漲幅 {rise_now:.2f}%"))
                    
                eligible = []
                for sym in filtered_stocks:
                    if sym == leader:
                        continue
                    # 2) 成交量突破
                    if not _vol_break(sym, first_condition_one_time):
                        continue
                    # 3) rise 先見高點後不再創高
                    if not _rise_peak_flat(sym, first_condition_one_time):
                        continue
                    # 4) 當下的漲幅、價格檢查
                    rise_now = row.get(f'rise_{sym}')
                    if rise_now is None or not (-2 <= rise_now <= 6):
                        continue
                    price_now = row.get(f'close_{sym}')
                    if price_now is None or price_now > capital_per_stock*1.5:
                        continue
                    row_sym = stock_data_collection[sym].loc[
                        stock_data_collection[sym]['time'] == current_time
                    ].iloc[0]
                    eligible.append({'symbol': sym, 'rise': rise_now, 'row': row_sym})

                if not eligible:
                    # 流程重置
                    pull_up_entry = limit_up_entry = False
                    tracking_stocks.clear()
                    if verbose:
                        message_log.append(
                            (current_time_str,
                             "等待結束無符合股票，流程重置"))
                else:
                    eligible.sort(key=lambda x: x['rise'], reverse=True)
                    chosen = eligible[len(eligible)//2]

                    # 進場與停損計算與 live 版一致
                    rowch   = chosen['row']
                    entry_p = rowch['close']
                    shares  = round((capital_per_stock*10000)/(entry_p*1000))
                    sell_cost = shares * entry_p * 1000
                    entry_fee = int(sell_cost * (transaction_fee*0.01) *
                                    (transaction_discount*0.01))
                    tax   = int(sell_cost * (trading_tax*0.01))
                    if entry_p < 10:
                        gap, tick = price_gap_below_50, 0.01
                    elif entry_p < 50:
                        gap, tick = price_gap_50_to_100, 0.05
                    elif entry_p < 100:
                        gap, tick = price_gap_50_to_100, 0.1
                    elif entry_p < 500:
                        gap, tick = price_gap_100_to_500, 0.5
                    elif entry_p < 1000:
                        gap, tick = price_gap_500_to_1000, 1
                    else:
                        gap, tick = price_gap_above_1000, 5

                    highest_on_entry = rowch['highest'] or entry_p
                    if (highest_on_entry-entry_p)*1000 < gap:
                        stop_thr = entry_p + gap/1000
                    else:
                        stop_thr = highest_on_entry + tick

                    current_position = {
                        'symbol': chosen['symbol'], 'shares': shares,
                        'entry_price': entry_p, 'sell_cost': sell_cost,
                        'entry_fee': entry_fee, 'tax': tax,
                        'entry_time': current_time_str,
                        'current_price_gap': gap, 'tick_unit': tick,
                        'highest_on_entry': highest_on_entry,
                        'stop_loss_threshold': stop_thr
                    }
                    in_position = True
                    has_exited  = False
                    hold_time   = 0
                    pull_up_entry = limit_up_entry = False
                    tracking_stocks.clear()
                    if verbose:
                        message_log.append(
                            (current_time_str,
                             f"{Fore.GREEN}進場！{chosen['symbol']} {shares}張 "
                             f"價 {entry_p:.2f} 停損 {stop_thr:.2f}"
                             f"{Style.RESET_ALL}"))
            else:
                # 🔁 新增：檢查是否再創新高
                if leader:
                    rise_now = row.get(f"rise_{leader}")
                    if leader_rise_before_decline is not None and rise_now is not None and rise_now > leader_rise_before_decline:
                        if verbose:
                            message_log.append(
                                (current_time_str,
                                f"{Fore.YELLOW}🚀 領漲股 {leader} 再創新高 {rise_now:.2f}% > {leader_rise_before_decline:.2f}%，觸發自我替換{Style.RESET_ALL}"))

                        leader_rise_before_decline = rise_now

                        # ✅ 終止等待（重點）
                        in_waiting_period = False
                        waiting_time = 0

                        if verbose:
                            message_log.append(
                                (current_time_str,
                                f"{Fore.MAGENTA}🚀 自我替換觸發，終止等待，重新觀察領漲{Style.RESET_ALL}"))
                        continue  # ⚡ 重點！不要繼續加 waiting_time

                # 如果沒有自我替換，才進行原本等待累加
                waiting_time += 1
                if verbose:
                    message_log.append(
                        (current_time_str,
                        f"等待中，第 {waiting_time} 分鐘"))

    # ═══════════ 2. 回測結果輸出 ═══════════ #
    message_log.sort(key=lambda x: x[0])
    for t, msg in message_log:
        print(f"[{t}] {msg}")

    if total_trades:
        avg_rate = total_profit_rate / total_trades
        print(f"\n模擬完成，總利潤：{int(total_profit)} 元，平均報酬率：{avg_rate:.2f}%\n")
        return total_profit, avg_rate
    else:
        print("無交易，無法計算利潤")
        return None, None
    
def exit_trade(
    selected_stock_df, shares, entry_price, sell_cost,
    entry_fee, tax,
    message_log, current_time, hold_time, entry_time, use_f_exit=False
):
    global transaction_fee, transaction_discount, trading_tax
    global in_position, has_exited, current_position
    current_time_str = current_time if isinstance(current_time, str) else current_time.strftime('%H:%M:%S')
    selected_stock_df['time'] = pd.to_datetime(selected_stock_df['time'], format='%H:%M:%S').dt.time

    if isinstance(entry_time, str):
        entry_time_obj = datetime.strptime(entry_time, '%H:%M:%S').time()
    else:
        entry_time_obj = entry_time

    if use_f_exit:
        end_time = datetime.strptime('13:30', '%H:%M').time()
        end_price_series = selected_stock_df[selected_stock_df['time'] == end_time]['close']
        if not end_price_series.empty:
            end_price = end_price_series.values[0]
        else:
            print("無法取得 13:30 的數據，出場時間配對錯誤")
            message_log.append((current_time_str, f"{RED}出場時間配對錯誤{RESET}"))
            return None, None
        entry_datetime = datetime.combine(date.today(), entry_time_obj)
        if isinstance(current_time, datetime):
            current_datetime = current_time
        else:
            current_datetime = datetime.combine(date.today(), current_time)
        hold_time_calculated = int((current_datetime - entry_datetime).total_seconds() / 60)
    else:
        entry_index_series = selected_stock_df[selected_stock_df['time'] == entry_time_obj].index
        if not entry_index_series.empty:
            entry_index = entry_index_series[0]
            exit_index = entry_index + hold_time
            if exit_index >= len(selected_stock_df):
                print("出場時間超出範圍，無法進行交易")
                message_log.append((current_time_str, f"{RED}出場時間超出範圍{RESET}"))
                return None, None
            end_price = selected_stock_df.iloc[exit_index]['close']
        else:
            print("進場時間配對錯誤，無法找到精確的進場時間")
            message_log.append((current_time_str, f"{RED}進場時間配對錯誤{RESET}"))
            return None, None
        hold_time_calculated = hold_time

    buy_cost = shares * end_price * 1000
    exit_fee = int(buy_cost * (transaction_fee * 0.01) * (transaction_discount * 0.01))
    profit = sell_cost - buy_cost - entry_fee - exit_fee - tax
    return_rate = (profit * 100) / (buy_cost - exit_fee) if (buy_cost - exit_fee) != 0 else 0.0

    if use_f_exit:
        message_log.append(
            (current_time_str,
             f"{RED}股票出場，持有時間 {hold_time_calculated} 分鐘（強制出場）{RESET}")
        )
    else:
        message_log.append(
            (current_time_str,
             f"{RED}股票出場，持有時間 {hold_time_calculated} 分鐘{RESET}")
        )
    message_log.append(
        (current_time_str,
         f"{RED}持有張數：{shares} 張，出場價格：{end_price} 元，出場價金：{int(buy_cost)} 元，利潤：{int(profit)} 元，"
         f"報酬率：{return_rate:.2f}%，手續費：{exit_fee} 元{RESET}")
    )

    in_position = False
    has_exited = True
    return profit, return_rate


# ------------------ 交易程式主程式 ------------------
#新增管理套件
to = tp.TouchOrderExecutor(api)

def process_live_trading_logic(
    symbols_to_analyze,
    current_time_str,
    wait_minutes,
    hold_minutes,
    message_log,
    in_position,
    has_exited,
    current_position,
    hold_time,
    already_entered_stocks,
    stop_loss_triggered,
    final_check_active,
    final_check_count,
    in_waiting_period,
    waiting_time,
    leader,
    tracking_stocks,
    previous_rise_values,
    leader_peak_rise,
    leader_rise_before_decline,
    first_condition_one_time,
    can_trade,
    group_positions,
    nb_matrix_path="nb_matrix_dict.json"
):
    """
    ────────────────────────────────────────────────────────────────────────
      盤中進場邏輯（漲停進場 / 拉高進場）
    ────────────────────────────────────────────────────────────────────────
    1. 觸發條件  
       ▸ 漲停進場：最新 K 棒 high == 漲停價，且前一 K 棒 high < 漲停價  
       ▸ 拉高進場：5 min 漲幅 ≥ 2% 且 volume > 1.5×(09:00~09:02 平均量)

    2. 追蹤清單（本版規則）  
       ─ 加入條件：5 min 漲幅 ≥ 1.5 %  
       ─ 加入時記錄 join_time、base_vol、base_rise

    3. 等待完成後的進場篩選  
       ❶ 非領漲  
       ❷ 自加入追蹤後 volume ≥ 1.5×(09:00~09:02 平均量) 曾出現  
       ❸ 自加入追蹤後 rise 先見高點且之後未再創高  
       ❹ 等待期滿當下 rise ∈ [-2 %, 6 %]

       → 依 rise 由大到小排序，取中間偏後股票下單  
         (市價 IOC 賣出 *day‑trade short*，TouchPrice 加停損買回)

    4. 其他流程（領漲偵測 / 反轉等待 / 最後十次檢查 / 停損計算）  
       沿用舊版，僅將涉及追蹤清單 & 進場挑選部分依新規則改寫。
    ────────────────────────────────────────────────────────────────────────
    """
    # ------------------------------ 0. 前置 ------------------------------- #
    monitor_stop_loss_orders()  # 偵測停損觸價單是否消失

    global capital_per_stock, transaction_fee, transaction_discount, trading_tax
    global price_gap_below_50, price_gap_50_to_100, price_gap_100_to_500
    global price_gap_500_to_1000, price_gap_above_1000, triggered_limit_up_stocks
    
    # Q 退出選單檢查
    if quit_flag['quit']:
        threading.Thread(target=show_exit_menu, daemon=True).start()
        quit_flag['quit'] = False

    # 時間解析
    try:
        current_dt = datetime.strptime(current_time_str, "%H:%M")
    except ValueError:
        print(f"時間格式錯誤：{current_time_str} (須 HH:MM)")
        return

    trading_time = current_dt.time()
    trading_txt  = current_dt.strftime("%H:%M:%S")

    # ---------- 讀 consolidated_symbols ----------
    if not os.path.exists(nb_matrix_path):
        print(f"找不到 {nb_matrix_path}")
        return
    with open(nb_matrix_path, "r", encoding="utf-8") as f:
        nb_dict = json.load(f)
    consolidated_symbols = nb_dict.get("consolidated_symbols", {})
    if not isinstance(consolidated_symbols, dict):
        print("consolidated_symbols 格式錯誤")
        return

    # ---------- 讀 auto_intraday ----------
    auto_intraday_file = "auto_intraday.json"
    if not os.path.exists(auto_intraday_file):
        print("缺少 auto_intraday.json")
        return
    with open(auto_intraday_file, "r", encoding="utf-8") as f:
        auto_intraday_data = json.load(f)

    # ---------- 建立 DataFrame 快取 ----------
    stock_df = {}
    for sym in symbols_to_analyze:
        data = auto_intraday_data.get(sym, [])
        df = pd.DataFrame(data).copy()
        if not df.empty:
            df["time"] = pd.to_datetime(df["time"], format="%H:%M:%S").dt.time
            df.sort_values("time", inplace=True)
            df.reset_index(drop=True, inplace=True)
        stock_df[sym] = df

    # ---------- 開盤前三分鐘均量 ----------
    FIRST3_AVG_VOL: dict[str, float] = {}
    for sym, df in stock_df.items():
        first3 = df[df["time"].astype(str).isin(["09:00:00", "09:01:00", "09:02:00"])]
        FIRST3_AVG_VOL[sym] = first3["volume"].mean() if not first3.empty else 0

    # ------------------------- 1. 觸發檢查 ------------------------------- #
    trigger_list = []

    # 13:00以後，不再觸發漲停或拉高進場
    if trading_time >= time(13, 0):
        print(f"⏰ {trading_txt} 已超過13:00，停止觸發新的漲停進場與拉高進場。")
    else:
        for grp, syms in consolidated_symbols.items():
            # 已經「觀察中」或「已進場」的族群不再檢查
            if grp in group_positions and group_positions[grp]:
                continue

            for sym in syms:
                if sym not in symbols_to_analyze:
                    continue
                df = stock_df[sym]
                if df.empty:
                    continue

                row_now = df[df["time"] == trading_time]
                if row_now.empty:
                    continue
                row_now = row_now.iloc[0]

                # ---- 漲停進場觸發 ----
                hit_limit = False
                if sym not in triggered_limit_up_stocks and row_now["high"] == row_now["漲停價"]:
                    prev_t = (datetime.combine(date.today(), trading_time) - timedelta(minutes=1)).time()
                    prev = df[df["time"] == prev_t]
                    prev_high = prev.iloc[0]["high"] if not prev.empty else None
                    if prev.empty or (prev_high is not None and prev_high < row_now["漲停價"]):
                        hit_limit = True
                        triggered_limit_up_stocks.add(sym)
                        print(f"[{trading_txt}] 股票 {sym} 第一次漲停觸發，加入已觸發清單")
                        # 取消同群拉高觀察
                        for g2, gstat in group_positions.items():
                            if isinstance(gstat, dict) and gstat.get("trigger") == "拉高進場":
                                if sym in consolidated_symbols.get(g2, []):
                                    group_positions[g2] = False
                                    msg = f"🚀 {sym} 漲停觸發！取消拉高觀察，轉由漲停進場處理"
                                    print(msg)
                                    message_log.append((trading_txt, msg))

                # ---- 拉高觸發 ----
                pull_up = False
                if row_now["2min_pct_increase"] >= 2:
                    avgv = FIRST3_AVG_VOL[sym]
                    if avgv and row_now["volume"] > 1.3 * avgv:
                        pull_up = True

                if hit_limit or pull_up:
                    trigger_list.append({
                        "symbol": sym,
                        "group": grp,
                        "condition": "limit_up" if hit_limit else "pull_up"
                    })

    trigger_list.sort(key=lambda x: 0 if x["condition"] == "limit_up" else 1)

    # ---------- 寫入觀察狀態 ----------
    for item in trigger_list:
        grp = item["group"]
        cond_txt = "漲停進場" if item["condition"] == "limit_up" else "拉高進場"
        if grp not in group_positions or not group_positions[grp]:
            group_positions[grp] = {
                "status": "觀察中",
                "trigger": cond_txt,
                "start_time": datetime.combine(date.today(), trading_time),
                "tracking": {},
                "leader": None
            }
            msg = f"族群 {grp} 進入觀察中（{cond_txt}）"
            print(msg)
            message_log.append((trading_txt, msg))

            if cond_txt == "漲停進場":
                gpos = group_positions[grp]
                gpos["wait_start"]   = datetime.combine(date.today(), trading_time)
                gpos["wait_counter"] = 0

    # ------------------------- 2. 更新追蹤清單 --------------------------- #
    for grp, gstat in group_positions.items():
        if not (isinstance(gstat, dict) and gstat["status"] == "觀察中"):
            continue

        # 確保 tracking dict 存在
        track = gstat.setdefault("tracking", {})

        for sym in consolidated_symbols[grp]:
            df = stock_df[sym]
            if df.empty:
                continue
            row_now = df[df["time"] == trading_time]
            if row_now.empty:
                continue
            row_now = row_now.iloc[0]

            # 加入條件：2min_pct_increase ≥ 2%
            if row_now["2min_pct_increase"] >= 2 and sym not in track:
                track[sym] = {
                    "join_time": datetime.combine(date.today(), trading_time),
                    "base_vol":  row_now["volume"],
                    "base_rise": row_now["rise"]
                }
                msg = f"{sym} 加入 {grp} 追蹤清單（2min↑2%）"
                print(msg)

    # ----------------------- 3. 領漲處理（拉高） ------------------------ #
    for grp, gstat in group_positions.items():
        if not (isinstance(gstat, dict) and gstat["status"] == "觀察中"):
            continue
        if gstat.get("trigger") != "拉高進場":
            continue

        track = gstat.get("tracking", {})
        if not track:
            continue

        # 目前 rise 最大者 = 領漲
        max_sym = None
        max_rise = None
        for sym in track:
            df = stock_df[sym]
            row_now = df[df["time"] == trading_time]
            if row_now.empty:
                continue
            rise_now = row_now.iloc[0]["rise"]
            if max_rise is None or rise_now > max_rise:
                max_rise = rise_now
                max_sym  = sym

        # 若首次確立領漲
        if gstat.get("leader") is None:
            gstat["leader"] = max_sym
            msg = f"拉高進場 {grp} 確立領漲：{max_sym}"
            print(msg)
            message_log.append((trading_txt, msg))
        else:
            # 若領漲替換
            if max_sym and max_sym != gstat["leader"]:
                print(f"領漲替換：{gstat['leader']} → {max_sym}")
                gstat["leader"] = max_sym
                gstat["leader_peak"] = rise_now
                gstat["leader_peak_time"] = current_time_str
                gstat["leader_reversal_rise"] = rise_now

                # ✅ 領漲替換後只更新，不進入等待
                gstat["status"] = "觀察中"
                gstat.pop("wait_start", None)
                gstat["wait_counter"] = 0
                print(f"🚀 領漲替換觸發，重新監控領漲")
        # ---- 領漲反轉 → 進入等待 ----
        lead_sym = gstat["leader"]
        if not lead_sym:
            continue
        df_lead = stock_df[lead_sym]
        idx_now = df_lead[df_lead["time"] == trading_time].index
        if idx_now.empty:
            continue
        idx_now = idx_now[0]
        if idx_now - 1 >= 0:
            high_now = df_lead.loc[idx_now, "high"]
            high_pre = df_lead.loc[idx_now - 1, "high"]
            if high_now <= high_pre:
                # 開始等待
                if "wait_start" not in gstat:
                    gstat["wait_start"] = now_full = datetime.combine(date.today(), trading_time)
                    gstat["wait_counter"] = 0
                    gstat["leader_reversal_rise"] = df_lead.loc[idx_now, "rise"]
                    msg = f"拉高進場 {grp} 領漲 {lead_sym} 反轉，開始等待"
                    print(msg)
                    message_log.append((trading_txt, msg))

    # --------- 若處於等待階段，每分鐘累加並印狀態 ---------
    for grp, gstat in group_positions.items():
        if not (isinstance(gstat, dict) and gstat["status"] == "觀察中"):
            continue

        # 漲停進場等待
        if gstat.get("trigger") == "漲停進場":
            if "wait_start" not in gstat:
                gstat["wait_start"]   = datetime.combine(date.today(), trading_time)
                gstat["wait_counter"] = 0
            else:
                gstat["wait_counter"] += 1
                print(f"漲停進場 {grp} 等待第 {gstat['wait_counter']} 分鐘")
            continue

        # 拉高進場等待（含自我替換檢查）
        if gstat.get("trigger") != "拉高進場" or "wait_start" not in gstat:
            continue

        # 自我替換檢查
        lead = gstat.get("leader")
        if lead and gstat.get("leader_reversal_rise") is not None:
            df_lead = stock_df.get(lead, pd.DataFrame())
            row_now = df_lead[df_lead["time"] == trading_time]
            if not row_now.empty:
                rise_now = row_now.iloc[0]["rise"]
                if rise_now > gstat["leader_reversal_rise"]:
                    print(f"🚀 領漲股 {lead} 再創新高 {rise_now:.2f}% > {gstat['leader_reversal_rise']:.2f}% ，觸發自我替換")
                    gstat["leader_reversal_rise"] = rise_now
                    # 回到觀察中，重置等待
                    gstat["status"]         = "觀察中"
                    gstat.pop("wait_start", None)
                    gstat["wait_counter"]   = 0
                    print("🚀 自我替換觸發，終止等待，重新觀察領漲")
                    continue

        # 正常累加等待分鐘
        gstat["wait_counter"] += 1
        print(f"拉高進場 {grp} 等待第 {gstat['wait_counter']} 分鐘")

    # ---------------- 4. 等待完成 → 篩選股票進場 ---------------- #
    def _vol_break(sym: str, join_time: datetime) -> bool:
        df = stock_df[sym]
        if df.empty:
            return False
        avgv = FIRST3_AVG_VOL[sym]
        if avgv == 0:
            return False
        later = df[df["time"] >= join_time.time()]
        return (later["volume"] >= 1.5 * avgv).any()

    def _rise_peak_flat(sym: str, join_time: datetime) -> bool:
        df = stock_df[sym]
        if df.empty:
            return False
        sub = df[df["time"] >= join_time.time()]
        if sub.empty:
            return False
        peak_idx = sub["rise"].idxmax()
        peak_val = sub.loc[peak_idx, "rise"]
        later = sub[sub.index > peak_idx]
        return (later["rise"] <= peak_val).all()

    groups_ready = []
    now_full = datetime.combine(date.today(), trading_time)
    for grp, gstat in group_positions.items():
        if not (isinstance(gstat, dict) and gstat["status"] == "觀察中"):
            continue
        elapsed = (now_full - gstat["start_time"]).total_seconds() / 60
        if elapsed >= wait_minutes:
            groups_ready.append(grp)

    for grp in groups_ready:
        gstat      = group_positions[grp]
        track      = gstat.get("tracking", {})   # 原始追蹤清單，dict[sym] = info
        leader_sym = gstat.get("leader")         # 領漲股票

        # —— 1. 排除 leader 自己 —— 
        candidates = [s for s in track.keys() if s != leader_sym]

        # —— 2. 盤中相似度篩選 —— 
        sim_ok = calculate_realtime_kline_data(
            leader=leader_sym,
            candidates=candidates,
            intraday_file="auto_intraday.json",
            wait_end_time=now_full,
            similarity_threshold=0.4,
            close_weight=2.0
        )
        print(f"[RT-SIM] 相似度合格：{sim_ok}")
        message_log.append((trading_txt, f"[RT-SIM] 相似度合格：{sim_ok}"))

        # —— 3. 濾除未通過相似度的股票 —— 
        filtered_track = {s: track[s] for s in sim_ok if s in track}
        if not filtered_track:
            msg = f"{grp} 相似度篩選後無候選 → 取消觀察"
            print(msg)
            message_log.append((trading_txt, msg))
            group_positions[grp] = False
            continue

        # —— 4. 更新追蹤清單 —— 
        gstat['tracking'] = filtered_track

        # 5. DEBUG：列出成交量突破門檻的股票
        passed_vol = []
        for s, info in filtered_track.items():
            df   = stock_df[s]
            join = info["join_time"]
            avgv = FIRST3_AVG_VOL.get(s, 0)
            later = df[df["time"] >= join.time()]
            ok = avgv > 0 and (later["volume"] >= 1.5 * avgv).any()
            passed_vol.append(s) if ok else None
            message_log.append(
                (trading_txt,
                 f"[DEBUG] {s} 開盤前三分鐘均量={avgv:.2f}，"
                 f"1.5倍門檻={1.5*avgv:.2f}，"
                 f"當前成交量={(later.iloc[-1]['volume'] if not later.empty else 0):.2f}，"
                 f"超越門檻={ok}")
            )

        # 6. DEBUG：列出反轉通過的股票與峰頂時間
        passed_rev = []
        for s in passed_vol:
            df = stock_df[s]
            sub = df[df["time"] >= filtered_track[s]["join_time"].time()]
            if sub.empty: continue
            pkidx = sub["rise"].idxmax()
            pkval = sub.loc[pkidx, "rise"]
            pk_time = sub.loc[pkidx, "time"].strftime("%H:%M:%S")
            later = sub.loc[pkidx+1:]
            later_max = later["rise"].max() if not later.empty else None
            ok = (later_max is None) or (later_max <= pkval)
            passed_rev.append((s, pk_time))
            message_log.append(
                (trading_txt,
                 f"[DEBUG] {s} 峰頂={pkval:.2f} at {pk_time}，"
                 f"後續最大={later_max}，pass={ok}")
            )

        # —— 7. 原本的 eligible 篩選邏輯 —— 
        eligible = []

        for sym, info in track.items():
            if sym == leader_sym:
                continue
            if not _vol_break(sym, info["join_time"]):
                continue
            if not _rise_peak_flat(sym, info["join_time"]):
                continue
            df = stock_df[sym]
            row_now = df[df["time"] == trading_time]
            if row_now.empty:
                continue
            rise_now = row_now.iloc[0]["rise"]
            if not (-2 <= rise_now <= 6):
                continue
            entry_price = row_now.iloc[0]["close"]
            if entry_price > capital_per_stock * 1.5:
                msg = f"⚠️ 排除 {sym}，股價 {entry_price:.2f} 超過資金上限 {capital_per_stock*1.5:.2f}"
                print(msg)
                message_log.append((trading_txt, msg))
                continue

            # ✅ 檢查是否可當沖
            try:
                stock_code_int = int(sym)
                contract = api.Contracts.Stocks[sym]  # Shioaji 內建合約查詢
                can_daytrade = contract.day_trade.value == "Yes"
                print(f"🧾 檢查 {sym} 是否可當沖：{contract.day_trade}")
                if not can_daytrade:
                    msg = f"⚠️ 排除 {sym}，今日不可當沖"
                    print(msg)
                    message_log.append((trading_txt, msg))
                    continue
            except Exception as e:
                msg = f"⚠️ 排除 {sym}，查詢 day_trade 欄位發生錯誤：{e}"
                print(msg)
                message_log.append((trading_txt, msg))
                continue

            eligible.append({
                "symbol": sym,
                "rise": rise_now,
                "row": row_now.iloc[0]
            })

        if not eligible:
            msg = f"{grp} 等待完成，但無符合條件股票 → 取消觀察"
            print(msg)
            message_log.append((trading_txt, msg))
            group_positions[grp] = False
            continue

        eligible.sort(key=lambda x: x["rise"], reverse=True)
        chosen = eligible[len(eligible)//2]

        # ------------------- 下單 -------------------
        
        row      = chosen["row"]
        entry_px = row["close"]
        shares   = round((capital_per_stock * 10000) / (entry_px * 1000))
        sell_amt = shares * entry_px * 1000
        fee      = int(sell_amt * (transaction_fee * 0.01) * (transaction_discount * 0.01))
        tax      = int(sell_amt * (trading_tax * 0.01))

        # --- 計算 gap & tick（依據 entry_price） ---
        if entry_px < 10:
            gap, tick = price_gap_below_50, 0.01
        elif entry_px < 50:
            gap, tick = price_gap_50_to_100, 0.05
        elif entry_px < 100:
            gap, tick = price_gap_50_to_100, 0.1
        elif entry_px < 500:
            gap, tick = price_gap_100_to_500, 0.5
        elif entry_px < 1000:
            gap, tick = price_gap_500_to_1000, 1
        else:
            gap, tick = price_gap_above_1000, 5

        # --- 原始停損價計算 ---
        highest_on_entry = row["highest"] or entry_px
        if (highest_on_entry - entry_px) * 1000 < gap:
            stop_type = "price_difference"
            stop_thr  = entry_px + gap / 1000
        else:
            stop_type = "over_high"
            stop_thr  = highest_on_entry + tick

        # ✅ 停損天花板限制：不得高於「漲停價 - 2 tick」
        limit_up = row["漲停價"]
        if limit_up < 10:
            tick_for_limit = 0.01
        elif limit_up < 50:
            tick_for_limit = 0.05
        elif limit_up < 100:
            tick_for_limit = 0.1
        elif limit_up < 500:
            tick_for_limit = 0.5
        elif limit_up < 1000:
            tick_for_limit = 1
        else:
            tick_for_limit = 5

        ceiling = limit_up - 2 * tick_for_limit
        if stop_thr > ceiling:
            stop_thr = ceiling
            stop_type = "ceiling_limit"
            print(f"🚫 停損價過高，啟用 ceiling 限制：{ceiling:.2f}（漲停 {limit_up} - 2 tick）")

        current_position = {
            "symbol": chosen["symbol"],
            "shares": shares,
            "entry_price": entry_px,
            "sell_cost": sell_amt,
            "entry_fee": fee,
            "tax": tax,
            "entry_time": trading_txt,
            "current_price_gap": gap,
            "tick_unit": tick,
            "highest_on_entry": highest_on_entry,
            "stop_loss_type": stop_type,
            "stop_loss_threshold": stop_thr
        }

        open_positions[chosen['symbol']] = {'entry_price': entry_px, 'shares': shares} # －將有真正送出單的股票加入到表中

        # --- 下市價 IOC 賣出單（券先賣） ---
        stock_code_int = int(chosen["symbol"])

        # 載入市場別映射表
        if not os.path.exists("twse_stocks_by_market.json"):
            print(f"⚠️ 找不到 twse_stocks_by_market.json，無法判別股票市場！取消下單。")
            return

        with open("twse_stocks_by_market.json", "r", encoding="utf-8") as f:
            stock_market_map = json.load(f)

        if chosen["symbol"] in stock_market_map.get("TSE", {}):
            contract = getattr(api.Contracts.Stocks.TSE, "TSE" + str(stock_code_int))
            market_label = "TSE"
        elif chosen["symbol"] in stock_market_map.get("OTC", {}):
            contract = getattr(api.Contracts.Stocks.OTC, "OTC" + str(stock_code_int))
            market_label = "OTC"
        else:
            print(f'⚠️ 找不到股票 {chosen["symbol"]} 的市場資料！取消下單。')
            return

        order = api.Order(
            price=0,
            quantity=shares,
            action=sj.constant.Action.Sell,
            price_type=sj.constant.StockPriceType.MKT,
            order_type=sj.constant.OrderType.IOC,
            order_lot=sj.constant.StockOrderLot.Common,
            daytrade_short=True,
            account=api.stock_account
        )
        trade = api.place_order(contract, order)

        # --- TouchPrice 停損買回 ---
        t_cmd = tp.TouchCmd(
            code=f"{stock_code_int}",
            close=tp.Price(price=stop_thr, trend="Equal")
        )
        o_cmd = tp.OrderCmd(
            code=f"{stock_code_int}",
            order=sj.Order(
                price=0,
                quantity=shares,
                action="Buy",
                order_type="ROD",
                price_type="MKT"
            )
        )
        tcond = tp.TouchOrderCond(t_cmd, o_cmd)
        to.add_condition(tcond)

        msg = (
            f"{GREEN}進場！{chosen['symbol']}  {shares}張  "
            f"成交價 {entry_px:.2f}  停損價 {stop_thr:.2f}{RESET}"
        )
        print(msg)
        message_log.append((trading_txt, msg))

        in_position            = True
        group_positions[grp]   = "已進場"
        leader                 = None
        tracking_stocks.clear()
        previous_rise_values.clear()

    # ------------------ 5. 依時間排序列印訊息 ------------------- #
    message_log.sort(key=lambda x: x[0])
    for t, m in message_log:
        print(f"[{t}] {m}")
    message_log.clear()

#盤中13:30出場
def exit_trade_live():
    """
    此函數依據設定，於 13:26 時進行出場動作：
      1. 從全域變數 to 中取得所有尚存的觸價委託單（to.conditions）
      2. 依據每個股票代號的所有委託單，累加取出進場張數（quantity），形成 exit_data 字典
      3. 將 exit_data 寫入本地檔案 "enter_exit.json"
      4. 重新讀取 "enter_exit.json" 的資料
      5. 對 exit_data 中每一筆資料，利用股票代號與進場張數建立出場委託單並下單
      6. 刪除所有尚存的觸價委託單
      7. 同步從 open_positions 中移除已平倉的股票
    """
    global open_positions

    # 1. 取得所有尚存的觸價委託單
    conditions_dict = to.conditions
    exit_data = {}

    # 2. 遍歷每個股票代號及其委託單列表，累加進場張數
    for stock_code, cond_list in conditions_dict.items():
        total_quantity = 0
        for cond in cond_list:
            try:
                qty = getattr(cond.order, 'quantity', 0)
                total_quantity += int(qty)
            except Exception as e:
                print(f"讀取股票 {stock_code} 的數量時發生錯誤：{e}")
        if total_quantity > 0:
            exit_data[stock_code] = total_quantity

    # 3. 將 exit_data 寫入 "enter_exit.json"
    try:
        with open("enter_exit.json", "w", encoding="utf-8") as f:
            json.dump(exit_data, f, ensure_ascii=False, indent=4)
        print("已將當前觸價委託單的股票代號和進場張數儲存至 enter_exit.json:")
        print(exit_data)
    except Exception as e:
        print(f"寫入 enter_exit.json 檔案失敗：{e}")
        return

    # 4. 讀取最新的 exit data
    try:
        with open("enter_exit.json", "r", encoding="utf-8") as f:
            exit_info = json.load(f)
    except Exception as e:
        print(f"讀取 enter_exit.json 檔案失敗：{e}")
        return

    if not exit_info:
        print("enter_exit.json 中沒有觸價委託單資料，終止出場程序。")
        return

    # 5. 對每筆 exit_info 中的資料，建立出場委託單並下單
    for stock_code, shares in exit_info.items():
        try:
            # 取得 contract 物件，例如 "TSE2330"
            contract = getattr(api.Contracts.Stocks.TSE, "TSE" + str(stock_code))
            limit_up_price = contract.limit_up

            # 建立限價買進的委託單 (ROC 條件)
            order = api.Order(
                action=sj.constant.Action.Buy,
                price=limit_up_price,
                quantity=shares,
                price_type=sj.constant.StockPriceType.LMT,
                order_type=sj.constant.OrderType.ROC,
                order_lot=sj.constant.StockOrderLot.Common,
                account=api.stock_account
            )
            trade = api.place_order(contract, order)
            print(f"{RED}下單出場：股票 {stock_code}，數量 {shares} 張；價格設定為漲停價 {limit_up_price}{RESET}")

            # 7. 同步從 open_positions 移除已平倉的股票
            open_positions.pop(stock_code, None)

        except Exception as e:
            print(f"處理股票 {stock_code} 時發生錯誤：{e}")

    # 6. 刪除所有尚存的觸價委託單
    for stock_code, cond_list in list(conditions_dict.items()):
        for cond in cond_list:
            try:
                to.delete_condition(cond)
            except Exception as e:
                print(f"刪除股票 {stock_code} 的觸價委託單時發生錯誤：{e}")

    print(f"{RED}出場委託單已全部下單，並刪除所有觸價委託單。{RESET}")

def list_open_positions():
    if not open_positions:
        print(f"{YELLOW}目前沒有任何持倉{RESET}")
        return
    print("\n========== 目前持倉 ==========")
    for i, (c, info) in enumerate(open_positions.items(), 1):
        print(f"{i}. {c:<6} {get_stock_name(c):<8} 進場價={info['entry_price']}  張數={info['shares']}")
    print("=" * 29)

def close_one_stock(code: str):
    """刪該股所有觸價單 + 以漲停價 ROC 市價買回"""
    conds = to.conditions.get(code, [])
    qty   = sum(getattr(c.order, 'quantity', 0) for c in conds)
    if qty == 0:
        print(f"⚠️  {code} 已無委託 / 持倉")
        return
    try:
        contract = getattr(api.Contracts.Stocks.TSE, f"TSE{code}")
        api.place_order(contract, api.Order(
            action=sj.constant.Action.Buy,
            price=contract.limit_up,
            quantity=qty,
            price_type=sj.constant.StockPriceType.LMT,
            order_type=sj.constant.OrderType.ROC,
            order_lot=sj.constant.StockOrderLot.Common,
            account=api.stock_account
        ))
        print(f"{GREEN}已平倉 {code}  共 {qty} 張{RESET}")
    except Exception as e:
        print(f"平倉 {code} 時錯誤：{e}")
    for c in conds:
        to.delete_condition(c)
    to.conditions.pop(code, None)
    open_positions.pop(code, None)

def main():
    load_settings()
    config = load_config("config.yaml")
    client = RestClient(api_key=config['api_key'])
    matrix_dict_analysis = load_matrix_dict_analysis()
    main_menu()

if __name__ == "__main__":
    '''
    #測試中文股票名稱
    data = fetch_twse_stock_codes(save_json="twse_stocks.json",
                                  save_csv="twse_stocks.csv")
    for code, name in data[:20]:
        print(code, name)
    '''
    ensure_packages(REQUIRED)
    print("開始執行程式...")
    main()